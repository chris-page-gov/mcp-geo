from __future__ import annotations

import json
import sqlite3
import subprocess
import sys
import zipfile
from io import BytesIO
from pathlib import Path

import scripts.ons_geo_cache_refresh as refresh
from openpyxl import Workbook

FIXTURES = Path(__file__).resolve().parent / "fixtures" / "ons_geo"


class _FakeResponse:
    def __init__(
        self,
        *,
        json_data=None,
        text: str | None = None,
        content: bytes | None = None,
        headers: dict[str, str] | None = None,
    ):
        self._json_data = json_data
        self.text = text or ""
        self.content = content if content is not None else self.text.encode("utf-8")
        content_type = "application/json" if json_data is not None else "text/html"
        self.headers = {"content-type": content_type, **(headers or {})}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def raise_for_status(self) -> None:
        return None

    def json(self):
        return self._json_data

    def iter_content(self, chunk_size: int = 1024 * 1024):
        del chunk_size
        yield self.content


def _write_manifest(tmp_path: Path, payload: dict) -> Path:
    path = tmp_path / "ons_geo_sources.json"
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
    return path


def _static_resolver(path: Path) -> dict[str, object]:
    return {"type": "static_file", "path": str(path)}


def _base_manifest() -> dict[str, object]:
    return {
        "version": "2026-04-08",
        "products": [
            {
                "id": "ONSPD",
                "title": "ONSPD",
                "keyType": "postcode",
                "derivationMode": "exact",
                "priority": 10,
                "release": "2026-02",
                "resolver": _static_resolver(FIXTURES / "onspd_modern.csv"),
                "semanticFields": {
                    "required": ["postcode", "lad_code"],
                    "optional": ["ward_code", "country_code", "region_code"],
                    "aliases": {"postcode": ["pcds"]},
                },
            },
            {
                "id": "NSPL",
                "title": "NSPL",
                "keyType": "postcode",
                "derivationMode": "best_fit",
                "priority": 20,
                "release": "2026-02",
                "resolver": _static_resolver(FIXTURES / "nspl_legacy.csv"),
                "semanticFields": {
                    "required": ["postcode", "lad_code"],
                    "optional": ["ward_code", "country_code", "region_code"],
                    "aliases": {"postcode": ["pcds"]},
                },
            },
            {
                "id": "ONSUD",
                "title": "ONSUD",
                "keyType": "uprn",
                "derivationMode": "exact",
                "priority": 10,
                "release": "2025-12",
                "resolver": _static_resolver(FIXTURES / "onsud_sample.csv"),
                "semanticFields": {
                    "required": ["uprn", "lad_code"],
                    "optional": [
                        "postcode",
                        "oa_code",
                        "lsoa_code",
                        "msoa_code",
                        "ward_code",
                        "country_code",
                        "region_code",
                        "postal_delivery",
                    ],
                    "aliases": {"uprn": ["UPRN"]},
                },
            },
            {
                "id": "NSUL",
                "title": "NSUL",
                "keyType": "uprn",
                "derivationMode": "best_fit",
                "priority": 20,
                "release": "2025-11",
                "resolver": _static_resolver(FIXTURES / "nsul_sample.csv"),
                "semanticFields": {
                    "required": ["uprn", "lad_code"],
                    "optional": [
                        "postcode",
                        "oa_code",
                        "lsoa_code",
                        "msoa_code",
                        "ward_code",
                        "country_code",
                        "region_code",
                        "postal_delivery",
                    ],
                    "aliases": {"uprn": ["UPRN"]},
                },
            },
        ],
        "supportProducts": [
            {
                "id": "CHD",
                "title": "CHD",
                "priority": 10,
                "release": "2025-12",
                "resolver": _static_resolver(FIXTURES / "chd_sample.csv"),
                "semanticFields": {
                    "required": ["code", "status"],
                    "optional": [
                        "name",
                        "successor_code",
                        "successor_name",
                        "code_family",
                        "level",
                    ],
                    "aliases": {
                        "code": ["GEOGRAPHY_CODE"],
                        "name": ["GEOGRAPHY_NAME"],
                        "status": ["STATUS"],
                        "successor_code": ["SUCCESSOR_CODE"],
                        "successor_name": ["SUCCESSOR_NAME"],
                        "code_family": ["CODE_FAMILY"],
                        "level": ["LEVEL"],
                    },
                    "defaults": {"status": "retired"},
                },
            },
            {
                "id": "RGC",
                "title": "RGC",
                "priority": 20,
                "release": "2025-12",
                "resolver": _static_resolver(FIXTURES / "rgc_current_sample.csv"),
                "semanticFields": {
                    "required": ["code", "name"],
                    "optional": ["status", "code_family", "level"],
                    "aliases": {
                        "code": ["GEOGRAPHY_CODE"],
                        "name": ["GEOGRAPHY_NAME"],
                        "status": ["STATUS"],
                        "code_family": ["CODE_FAMILY"],
                        "level": ["LEVEL"],
                    },
                    "defaults": {"status": "current"},
                },
            },
        ],
    }


def test_ons_geo_cache_refresh_ingests_products_and_sidecars(tmp_path: Path) -> None:
    manifest = _write_manifest(tmp_path, _base_manifest())
    cache_dir = tmp_path / "cache"
    index_path = tmp_path / "ons_geo_cache_index.json"
    db_name = "ons_geo_cache.sqlite"

    result = subprocess.run(
        [
            sys.executable,
            "scripts/ons_geo_cache_refresh.py",
            "--sources",
            str(manifest),
            "--cache-dir",
            str(cache_dir),
            "--index-path",
            str(index_path),
            "--db-name",
            db_name,
        ],
        capture_output=True,
        text=True,
        check=False,
        cwd=Path(__file__).resolve().parent.parent,
    )
    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert payload["health"]["status"] == "ready"
    assert payload["health"]["exactReady"] is True
    assert payload["health"]["bestFitReady"] is True
    assert payload["health"]["supportReady"] is True
    assert len(payload["supportProducts"]) == 2
    assert len(payload["products"]) == 4
    assert all(item["status"] == "ingested" for item in payload["supportProducts"])
    assert all(item["status"] == "ingested" for item in payload["products"])
    assert payload["products"][0]["schemaValidation"]["schemaFingerprint"]

    conn = sqlite3.connect(str(cache_dir / db_name))
    row_count = conn.execute("SELECT COUNT(*) FROM ons_geo_rows").fetchone()[0]
    code_count = conn.execute("SELECT COUNT(*) FROM ons_geo_code_reference").fetchone()[0]
    uprn_index_count = conn.execute("SELECT COUNT(*) FROM ons_geo_uprn_index").fetchone()[0]
    normalized = conn.execute(
        "SELECT normalized_json FROM ons_geo_rows "
        "WHERE product_id = 'ONSPD' AND key_norm = 'SW1A1AA'"
    ).fetchone()
    uprn_row = conn.execute(
        """
        SELECT lad_code, ward_code, country_code, region_code
        FROM ons_geo_uprn_index
        WHERE product_id = 'ONSUD' AND uprn = '100023336959'
        """
    ).fetchone()
    conn.close()

    assert row_count == 4
    assert code_count >= 6
    assert uprn_index_count == 2
    assert normalized is not None
    normalized_payload = json.loads(normalized[0])
    assert normalized_payload["geographies"]["lad"]["currentCode"] == "E09000033"
    assert normalized_payload["geographies"]["ward"]["currentCode"] == "E05013806"
    assert uprn_row == ("E08000026", "E05001111", "E92000001", "E12000005")


def test_ons_geo_cache_refresh_reports_partial_failure_for_missing_support_dataset(
    tmp_path: Path,
) -> None:
    manifest_payload = _base_manifest()
    manifest_payload["supportProducts"][1]["resolver"] = _static_resolver(
        tmp_path / "missing-rgc.csv"
    )
    manifest = _write_manifest(tmp_path, manifest_payload)
    cache_dir = tmp_path / "cache"
    index_path = tmp_path / "ons_geo_cache_index.json"

    result = subprocess.run(
        [
            sys.executable,
            "scripts/ons_geo_cache_refresh.py",
            "--sources",
            str(manifest),
            "--cache-dir",
            str(cache_dir),
            "--index-path",
            str(index_path),
        ],
        capture_output=True,
        text=True,
        check=False,
        cwd=Path(__file__).resolve().parent.parent,
    )
    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert payload["health"]["status"] == "degraded"
    assert "support_datasets_unavailable" in payload["health"]["degradedReasons"]
    rgc = next(item for item in payload["supportProducts"] if item["id"] == "RGC")
    assert rgc["status"] == "error"
    assert rgc["errorCode"] == "RESOLVE_ERROR"


def test_index_health_requires_each_key_type_per_mode() -> None:
    products = [
        {
            "kind": "product",
            "id": "ONSPD",
            "keyType": "postcode",
            "derivationMode": "exact",
            "status": "ingested",
        },
        {
            "kind": "product",
            "id": "ONSUD",
            "keyType": "uprn",
            "derivationMode": "exact",
            "status": "error",
        },
        {
            "kind": "product",
            "id": "NSPL",
            "keyType": "postcode",
            "derivationMode": "best_fit",
            "status": "ingested",
        },
        {
            "kind": "product",
            "id": "NSUL",
            "keyType": "uprn",
            "derivationMode": "best_fit",
            "status": "ingested",
        },
    ]

    health = refresh._index_health(products, [])

    assert health["exactReady"] is False
    assert health["bestFitReady"] is True
    assert health["status"] == "degraded"
    assert "exact_products_unavailable" in health["degradedReasons"]


def test_resolve_hosted_table_arcgis_pages_rows_and_extracts_schema(
    monkeypatch,
    tmp_path: Path,
) -> None:
    metadata = json.loads((FIXTURES / "onspd_arcgis_metadata.json").read_text(encoding="utf-8"))
    page = json.loads((FIXTURES / "onspd_arcgis_page_1.json").read_text(encoding="utf-8"))

    calls: list[tuple[str, dict[str, object]]] = []

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream
        calls.append((url, params or {}))
        if url.endswith("?f=json"):
            return _FakeResponse(json_data=metadata)
        return _FakeResponse(json_data=page)

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(_write_manifest(tmp_path, {
        "version": "2026-04-08",
        "products": [
            {
                "id": "ONSPD",
                "title": "ONSPD",
                "keyType": "postcode",
                "derivationMode": "exact",
                "priority": 10,
                "release": "latest",
                "resolver": {
                    "type": "hosted_table_arcgis",
                    "metadataUrl": "https://example.test/FeatureServer/0?f=json",
                    "queryUrl": "https://example.test/FeatureServer/0/query"
                },
                "semanticFields": {
                    "required": ["postcode", "lad_code"],
                    "optional": ["ward_code"],
                    "aliases": {"postcode": ["pcds"]}
                }
            }
        ],
        "supportProducts": []
    }))[1][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.source_format == "ndjson"
    assert resolved.schema_fields[:2] == ["OBJECTID", "pcds"]
    assert resolved.field_aliases["lad25cd"] == "Local Authority District Code (2025)"
    content = resolved.source_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(content) == 1
    assert json.loads(content[0])["pcds"] == "SW1A1AA"
    assert any(call[0].endswith("/query") for call in calls)


def test_resolve_hosted_table_arcgis_uses_objectid_keyset_pagination(
    monkeypatch,
    tmp_path: Path,
) -> None:
    metadata_url = "https://example.test/FeatureServer/0?f=json"
    query_url = "https://example.test/FeatureServer/0/query"
    metadata = {
        "name": "ArcGIS Hub Dataset",
        "maxRecordCount": 1,
        "objectIdField": "OBJECTID",
        "fields": [
            {"name": "OBJECTID", "alias": "Object ID"},
            {"name": "pcds", "alias": "Postcode"},
            {"name": "lad25cd", "alias": "Local authority district code"},
        ],
    }
    calls: list[tuple[str, dict[str, object]]] = []

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream
        calls.append((url, params or {}))
        if url == metadata_url:
            return _FakeResponse(json_data=metadata)
        assert params is not None
        where = params.get("where")
        if where == "1=1":
            return _FakeResponse(
                json_data={
                    "features": [
                        {
                            "attributes": {
                                "OBJECTID": 1,
                                "pcds": "SW1A1AA",
                                "lad25cd": "E09000033",
                            }
                        }
                    ],
                    "exceededTransferLimit": True,
                }
            )
        if where == "OBJECTID > 1":
            return _FakeResponse(
                json_data={
                    "features": [
                        {
                            "attributes": {
                                "OBJECTID": 2,
                                "pcds": "SW1A2AA",
                                "lad25cd": "E09000033",
                            }
                        }
                    ],
                    "exceededTransferLimit": False,
                }
            )
        return _FakeResponse(json_data={"features": [], "exceededTransferLimit": False})

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [
                    {
                        "id": "ONSPD",
                        "title": "ONSPD",
                        "keyType": "postcode",
                        "derivationMode": "exact",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "hosted_table_arcgis",
                            "metadataUrl": metadata_url,
                            "queryUrl": query_url,
                        },
                        "semanticFields": {
                            "required": ["postcode", "lad_code"],
                            "optional": [],
                            "aliases": {
                                "postcode": ["pcds"],
                                "lad_code": ["lad25cd"],
                            },
                        },
                    }
                ],
                "supportProducts": [],
            },
        )
    )[1][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )

    query_calls = [params for url, params in calls if url == query_url]
    assert len(query_calls) >= 2
    assert query_calls[0]["where"] == "1=1"
    assert query_calls[1]["where"] == "OBJECTID > 1"
    assert "resultOffset" not in query_calls[1]
    content = resolved.source_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(content) == 2


def test_resolve_hosted_table_arcgis_rejects_non_progressing_offsets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    metadata_url = "https://example.test/FeatureServer/0?f=json"
    query_url = "https://example.test/FeatureServer/0/query"
    repeated_page = {
        "features": [
            {"attributes": {"OBJECTID": 1, "UPRN": "100023336959", "LAD24CD": "E08000026"}},
            {"attributes": {"OBJECTID": 2, "UPRN": "100023336960", "LAD24CD": "E08000026"}},
        ],
        "exceededTransferLimit": True,
    }

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream
        if url == metadata_url:
            return _FakeResponse(
                json_data={
                    "name": "ArcGIS Hub Dataset",
                    "maxRecordCount": 2,
                    "objectIdField": "OBJECTID",
                    "fields": [
                        {"name": "OBJECTID", "alias": "Object ID"},
                        {"name": "UPRN", "alias": "UPRN"},
                        {"name": "LAD24CD", "alias": "Local authority district code"},
                    ],
                }
            )
        if url == query_url:
            assert params is not None
            return _FakeResponse(json_data=repeated_page)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(_write_manifest(tmp_path, {
        "version": "2026-04-08",
        "products": [
            {
                "id": "ONSUD",
                "title": "ONSUD",
                "keyType": "uprn",
                "derivationMode": "exact",
                "priority": 10,
                "release": "latest",
                "resolver": {
                    "type": "hosted_table_arcgis",
                    "metadataUrl": metadata_url,
                    "queryUrl": query_url
                },
                "semanticFields": {
                    "required": ["uprn", "lad_code"],
                    "optional": [],
                    "aliases": {"uprn": ["UPRN"], "lad_code": ["LAD24CD"]}
                }
            }
        ],
        "supportProducts": []
    }))[1][0]

    try:
        refresh.resolve_dataset_source(
            dataset,
            raw_root=tmp_path / "raw",
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "pagination made no progress" in str(exc)
    else:
        raise AssertionError("Expected non-progressing ArcGIS pagination to be rejected")


def test_ingest_support_dataset_accepts_rgc_xlsx_archive(tmp_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "RGC"
    summary.append(["Entity code", "Entity name"])
    summary.append(["E05", "Wards"])

    metadata = workbook.create_sheet("Metadata_for_geography_listings")
    metadata.append(["Attribute", "Description"])
    metadata.append(["GEOGCD", "Code"])

    ward = workbook.create_sheet("E05_WD")
    ward.append(["GEOGCD", "GEOGNM", "STATUS", "ENTITYCD"])
    ward.append(["E05000001", "Aldersgate", "live", "E05"])
    ward.append(["E05000002", "Aldgate", "terminated", "E05"])

    country = workbook.create_sheet("E92_CTRY")
    country.append(["GEOGCD", "GEOGNM", "STATUS", "ENTITYCD"])
    country.append(["E92000001", "England", "live", "E92"])

    xlsx_bytes = BytesIO()
    workbook.save(xlsx_bytes)
    workbook.close()

    archive_path = tmp_path / "rgc.zip"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("rgc.xlsx", xlsx_bytes.getvalue())

    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "RGC",
                        "title": "Register of Geographic Codes",
                        "priority": 20,
                        "release": "latest",
                        "resolver": _static_resolver(archive_path),
                        "semanticFields": {
                            "required": ["code", "name"],
                            "optional": ["status", "code_family", "level"],
                            "aliases": {
                                "code": ["GEOGRAPHY_CODE", "ENTITYCD"],
                                "name": ["GEOGRAPHY_NAME", "ENTITYNM"],
                                "status": ["STATUS"],
                                "code_family": ["CODE_FAMILY"],
                                "level": ["LEVEL"],
                            },
                            "defaults": {"status": "current"},
                        },
                    }
                ],
            },
        )
    )[2][0]

    resolved = refresh.ResolvedSource(
        dataset_id="RGC",
        source_path=archive_path,
        source_format="zip",
        resolved_source_url=None,
        resolved_release="December 2025",
        retrieved_at="2026-04-10T00:00:00Z",
        metadata={},
        schema_fields=[],
        field_aliases={},
    )
    conn = sqlite3.connect(":memory:")
    refresh.ensure_schema(conn)
    code_references = refresh.CodeReferenceStore()
    inserted, schema_validation = refresh._ingest_support_dataset(
        conn=conn,
        dataset=dataset,
        resolved=resolved,
        max_rows=None,
        code_references=code_references,
    )

    assert inserted == 3
    assert schema_validation["status"] == "ok"
    assert code_references.annotate("E05000001", "ward")["status"] == "current"
    assert code_references.annotate("E05000001", "ward")["codeFamily"] == "ward"
    assert code_references.annotate("E92000001", "country")["codeFamily"] == "country"
    assert conn.execute("SELECT COUNT(*) FROM ons_geo_code_reference").fetchone()[0] == 3
    conn.close()


def test_resolve_portal_release_file_prefers_discovery_api_and_suffix(
    monkeypatch,
    tmp_path: Path,
) -> None:
    package_show = json.loads((FIXTURES / "portal_package_show.json").read_text(encoding="utf-8"))
    landing_html = (FIXTURES / "portal_release_page.html").read_text(encoding="utf-8")
    zip_content = b"PK\x03\x04fake"

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/api/package_show":
            return _FakeResponse(json_data=package_show)
        if url == "https://example.test/landing":
            return _FakeResponse(text=landing_html)
        if url.endswith(".zip"):
            return _FakeResponse(content=zip_content)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(_write_manifest(tmp_path, {
        "version": "2026-04-08",
        "products": [],
        "supportProducts": [
            {
                "id": "ONSUD",
                "title": "ONSUD",
                "priority": 10,
                "release": "latest",
                "resolver": {
                    "type": "portal_release_file",
                    "landingUrl": "https://example.test/landing",
                    "discoveryApiUrl": "https://example.test/api/package_show",
                    "preferredSuffixes": [".zip"],
                    "linkPatterns": ["onsud", "zip"],
                    "releasePatterns": ["Epoch\\s+\\d+"]
                },
                "semanticFields": {
                    "required": ["code"],
                    "optional": [],
                    "aliases": {"code": ["GEOGRAPHY_CODE"]}
                }
            }
        ]
    }))[2][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.resolved_source_url == "https://example.test/downloads/onsud-december-2025-epoch-123.zip"
    assert resolved.resolved_release == "December 2025 (Epoch 123)"
    assert resolved.source_path.exists()


def test_resolve_portal_release_file_skips_failing_landing_when_discovery_has_zip(
    monkeypatch,
    tmp_path: Path,
) -> None:
    package_show = {
        "success": True,
        "result": {
            "name": "ons-uprn-directory-december-2025-epoch-123",
            "title": "ONS UPRN Directory (December 2025) (Epoch 123)",
            "resources": [
                {
                    "name": "ZIP download",
                    "format": "ZIP",
                    "url": "https://downloads.example.test/onsud-december-2025-epoch-123.zip",
                }
            ],
        },
    }
    zip_content = b"PK\x03\x04direct"

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/api/package_show":
            return _FakeResponse(json_data=package_show)
        if url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip":
            return _FakeResponse(content=zip_content)
        if url == "https://www.data.gov.uk/dataset/ons-uprn-directory-december-2025-epoch-123":
            raise AssertionError(
                "landing page should not be fetched when discovery already yields zip"
            )
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "discoveryApiUrl": "https://example.test/api/package_show",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "ons-uprn-directory", "zip"],
                            "releasePatterns": ["Epoch\\s+\\d+"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.resolved_source_url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip"
    assert resolved.resolved_release == "December 2025 (Epoch 123)"


def test_resolve_portal_release_file_uses_latest_ckan_search_result(
    monkeypatch,
    tmp_path: Path,
) -> None:
    search_payload = {
        "success": True,
        "result": {
            "results": [
                {
                    "name": "ons-uprn-directory-october-2025-epoch-121",
                    "title": "ONS UPRN Directory (October 2025) (Epoch 121)",
                    "metadata_modified": "2026-01-15T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-october-2025-epoch-121",
                        }
                    ],
                },
                {
                    "name": "ons-uprn-directory-january-2026-epoch-124",
                    "title": "ONS UPRN Directory (January 2026) (Epoch 124)",
                    "metadata_modified": "2025-12-20T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-january-2026-epoch-124",
                        }
                    ],
                },
                {
                    "name": "ons-uprn-directory-december-2025-epoch-123-user-guide",
                    "title": "ONS UPRN Directory (December 2025) (Epoch 123) User Guide",
                    "metadata_modified": "2025-12-24T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123-user-guide",
                        }
                    ],
                },
                {
                    "name": "ons-uprn-directory-december-2025-epoch-123",
                    "title": "ONS UPRN Directory (December 2025) (Epoch 123)",
                    "metadata_modified": "2025-12-23T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123",
                        }
                    ],
                },
            ]
        },
    }
    data_gov_landing = (
        '<html><body><a href="https://open-geography.example/datasets/'
        'ons::ons-uprn-directory-january-2026-epoch-124">hub</a></body></html>'
    )
    hub_html = (
        '<html><body><a href="https://downloads.example.test/onsud-january-2026-epoch-124.zip">'
        "download</a></body></html>"
    )
    zip_content = b"PK\x03\x04latest"

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/api/package_search":
            return _FakeResponse(json_data=search_payload)
        if url == "https://www.data.gov.uk/dataset/ons-uprn-directory-january-2026-epoch-124":
            return _FakeResponse(text=data_gov_landing)
        if url == "https://open-geography.example/datasets/ons::ons-uprn-directory-january-2026-epoch-124":
            return _FakeResponse(text=hub_html)
        if url == "https://downloads.example.test/onsud-january-2026-epoch-124.zip":
            return _FakeResponse(content=zip_content)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/latest",
                            "discoveryApiUrl": "https://example.test/api/package_search",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "ons-uprn-directory", "zip"],
                            "releasePatterns": [
                                "(January|February|March|April|May|June|July|August|September|"
                                "October|November|December)\\s+20\\d{2}",
                                "Epoch\\s+\\d+",
                            ],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.resolved_source_url == "https://downloads.example.test/onsud-january-2026-epoch-124.zip"
    assert resolved.resolved_release == "January 2026 (Epoch 124)"
    assert resolved.source_path.exists()


def test_resolve_portal_release_file_ignores_unrelated_newer_ckan_result(
    monkeypatch,
    tmp_path: Path,
) -> None:
    search_payload = {
        "success": True,
        "result": {
            "results": [
                {
                    "name": "register-of-geographic-codes-december-2025-for-the-uk",
                    "title": "Register of Geographic Codes (December 2025) for the UK",
                    "metadata_modified": "2026-04-09T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::register-of-geographic-codes-december-2025-for-the-uk",
                        }
                    ],
                },
                {
                    "name": "ons-uprn-directory-december-2025-epoch-123",
                    "title": "ONS UPRN Directory (December 2025) (Epoch 123)",
                    "metadata_modified": "2025-12-23T12:00:00.000000",
                    "resources": [
                        {
                            "name": "ArcGIS Hub Dataset",
                            "format": "HTML",
                            "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123",
                        }
                    ],
                },
            ]
        },
    }
    data_gov_landing = (
        '<html><body><a href="https://open-geography.example/datasets/'
        'ons::ons-uprn-directory-december-2025-epoch-123">hub</a></body></html>'
    )
    hub_html = (
        '<html><body><a href="https://downloads.example.test/onsud-december-2025-epoch-123.zip">'
        "download</a></body></html>"
    )
    zip_content = b"PK\x03\x04onsud"

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/api/package_search":
            return _FakeResponse(json_data=search_payload)
        if url == "https://www.data.gov.uk/dataset/ons-uprn-directory-december-2025-epoch-123":
            return _FakeResponse(text=data_gov_landing)
        if url == "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123":
            return _FakeResponse(text=hub_html)
        if url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip":
            return _FakeResponse(content=zip_content)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/latest",
                            "discoveryApiUrl": "https://example.test/api/package_search",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "ons-uprn-directory", "zip"],
                            "releasePatterns": [
                                "(January|February|March|April|May|June|July|August|September|"
                                "October|November|December)\\s+20\\d{2}",
                                "Epoch\\s+\\d+",
                            ],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.resolved_source_url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip"
    assert resolved.resolved_release == "December 2025 (Epoch 123)"


def test_resolve_portal_release_file_rejects_unsupported_binary_release_assets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    landing_html = (
        '<html><body><a href="https://downloads.example.test/chd-december-2025.xlsx">'
        "download</a></body></html>"
    )

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/landing":
            return _FakeResponse(text=landing_html)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "CHD",
                        "title": "CHD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "preferredSuffixes": [".zip", ".csv", ".xlsx", ".mdb"],
                            "linkPatterns": ["chd"],
                            "releasePatterns": ["December\\s+2025"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    try:
        refresh.resolve_dataset_source(
            dataset,
            raw_root=tmp_path / "raw",
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "unsupported format .xlsx" in str(exc)
    else:
        raise AssertionError("Expected unsupported XLSX release asset to be rejected")


def test_probe_portal_release_file_rejects_unsupported_binary_release_assets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    landing_html = (
        '<html><body><a href="https://downloads.example.test/chd-december-2025.mdb">'
        "download</a></body></html>"
    )

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/landing":
            return _FakeResponse(text=landing_html)
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "CHD",
                        "title": "CHD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "preferredSuffixes": [".zip", ".csv", ".xlsx", ".mdb"],
                            "linkPatterns": ["chd"],
                            "releasePatterns": ["December\\s+2025"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    try:
        refresh.probe_dataset_source(
            dataset,
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "unsupported format .mdb" in str(exc)
    else:
        raise AssertionError("Expected unsupported MDB release asset to be rejected")


def test_probe_portal_release_file_skips_failing_landing_when_discovery_has_zip(
    monkeypatch,
    tmp_path: Path,
) -> None:
    package_show = {
        "success": True,
        "result": {
            "name": "ons-uprn-directory-december-2025-epoch-123",
            "title": "ONS UPRN Directory (December 2025) (Epoch 123)",
            "resources": [
                {
                    "name": "ZIP download",
                    "format": "ZIP",
                    "url": "https://downloads.example.test/onsud-december-2025-epoch-123.zip",
                }
            ],
        },
    }

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, params
        if url == "https://example.test/api/package_show":
            return _FakeResponse(json_data=package_show)
        if url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip":
            return _FakeResponse(content=b"PK\x03\x04direct")
        if url == "https://www.data.gov.uk/dataset/ons-uprn-directory-december-2025-epoch-123":
            raise AssertionError(
                "landing page should not be fetched when discovery already yields zip"
            )
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "discoveryApiUrl": "https://example.test/api/package_show",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "ons-uprn-directory", "zip"],
                            "releasePatterns": ["Epoch\\s+\\d+"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    probe = refresh.probe_dataset_source(
        dataset,
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert probe.resolved_source_url == "https://downloads.example.test/onsud-december-2025-epoch-123.zip"
    assert probe.resolved_release == "December 2025 (Epoch 123)"


def test_resolve_portal_release_file_rejects_suffixless_release_assets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    landing_html = (
        '<html><body><a href="https://downloads.example.test/chd/download">'
        "download</a></body></html>"
    )

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/landing":
            return _FakeResponse(text=landing_html)
        if url == "https://downloads.example.test/chd/download":
            return _FakeResponse(text="<html><body>No file</body></html>")
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "CHD",
                        "title": "CHD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "preferredSuffixes": [".zip", ".csv"],
                            "linkPatterns": ["chd", "download"],
                            "releasePatterns": ["December\\s+2025"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    try:
        refresh.resolve_dataset_source(
            dataset,
            raw_root=tmp_path / "raw",
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "no ingestible file suffix" in str(exc)
    else:
        raise AssertionError("Expected suffixless release asset to be rejected")


def test_probe_portal_release_file_rejects_suffixless_release_assets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    landing_html = (
        '<html><body><a href="https://downloads.example.test/chd/download">'
        "download</a></body></html>"
    )

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://example.test/landing":
            return _FakeResponse(text=landing_html)
        if url == "https://downloads.example.test/chd/download":
            return _FakeResponse(text="<html><body>No file</body></html>")
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "CHD",
                        "title": "CHD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "preferredSuffixes": [".zip", ".csv"],
                            "linkPatterns": ["chd", "download"],
                            "releasePatterns": ["December\\s+2025"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    try:
        refresh.probe_dataset_source(
            dataset,
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "no ingestible file suffix" in str(exc)
    else:
        raise AssertionError("Expected suffixless release asset to be rejected")


def test_resolve_portal_release_file_uses_arcgis_guid_download_when_discovery_only_has_html(
    monkeypatch,
    tmp_path: Path,
) -> None:
    package_show = {
        "success": True,
        "result": {
            "name": "ons-uprn-directory-december-2025-epoch-123",
            "title": "ONS UPRN Directory (December 2025) (Epoch 123)",
            "extras": [
                {
                    "key": "guid",
                    "value": "https://www.arcgis.com/home/item.html?id=cf1e4c08e78d48e387bcfab837f4e1d0",
                }
            ],
            "resources": [
                {
                    "name": "ArcGIS Hub Dataset",
                    "format": "HTML",
                    "url": "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123",
                }
            ],
        },
    }
    arcgis_data_url = (
        "https://www.arcgis.com/sharing/rest/content/items/"
        "cf1e4c08e78d48e387bcfab837f4e1d0/data"
    )

    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, params
        if url == "https://example.test/api/package_show":
            return _FakeResponse(json_data=package_show)
        if url == "https://open-geography.example/datasets/ons::ons-uprn-directory-december-2025-epoch-123":
            return _FakeResponse(text="<html><body>hub page</body></html>")
        if url == arcgis_data_url:
            return _FakeResponse(
                content=b"PK\x03\x04guid",
                headers={
                    "content-type": "application/zip",
                    "content-disposition": 'attachment; filename="ONSUD_DEC_2025.zip"',
                },
            )
        if url == "https://www.data.gov.uk/dataset/ons-uprn-directory-december-2025-epoch-123":
            raise AssertionError(
                "landing page should not be fetched when ArcGIS guid yields direct download"
            )
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "discoveryApiUrl": "https://example.test/api/package_show",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "ons-uprn-directory", "zip"],
                            "releasePatterns": [
                                "(January|February|March|April|May|June|July|August|September|"
                                "October|November|December)\\s+20\\d{2}",
                                "Epoch\\s+\\d+",
                            ],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    resolved = refresh.resolve_dataset_source(
        dataset,
        raw_root=tmp_path / "raw",
        timeout=5.0,
        file_overrides={},
        url_overrides={},
    )
    assert resolved.resolved_source_url == arcgis_data_url
    assert resolved.resolved_release == "December 2025 (Epoch 123)"
    assert resolved.source_format == "zip"
    assert resolved.source_path.exists()


def test_resolve_direct_url_rejects_suffixless_urls(
    monkeypatch,
    tmp_path: Path,
) -> None:
    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://downloads.example.test/onsud/download":
            return _FakeResponse(text="<html><body>No file</body></html>")
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "direct_url",
                            "downloadUrl": "https://downloads.example.test/onsud/download",
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    try:
        refresh.resolve_dataset_source(
            dataset,
            raw_root=tmp_path / "raw",
            timeout=5.0,
            file_overrides={},
            url_overrides={},
        )
    except ValueError as exc:
        assert "Direct URL source has no ingestible file suffix" in str(exc)
    else:
        raise AssertionError("Expected suffixless direct URL to be rejected")


def test_probe_dataset_source_rejects_suffixless_url_overrides(
    monkeypatch,
    tmp_path: Path,
) -> None:
    def fake_get(url, timeout=None, stream=False, params=None):
        del timeout, stream, params
        if url == "https://downloads.example.test/onsud/download":
            return _FakeResponse(text="<html><body>No file</body></html>")
        raise AssertionError(f"Unexpected URL {url}")

    monkeypatch.setattr(refresh.requests, "get", fake_get)
    dataset = refresh.load_manifest(
        _write_manifest(
            tmp_path,
            {
                "version": "2026-04-08",
                "products": [],
                "supportProducts": [
                    {
                        "id": "ONSUD",
                        "title": "ONSUD",
                        "priority": 10,
                        "release": "latest",
                        "resolver": {
                            "type": "portal_release_file",
                            "landingUrl": "https://example.test/landing",
                            "preferredSuffixes": [".zip"],
                            "linkPatterns": ["onsud", "zip"],
                            "releasePatterns": ["Epoch\\s+\\d+"],
                        },
                        "semanticFields": {
                            "required": ["code"],
                            "optional": [],
                            "aliases": {"code": ["GEOGRAPHY_CODE"]},
                        },
                    }
                ],
            },
        )
    )[2][0]

    def fail_if_called(*args, **kwargs):
        del args, kwargs
        raise AssertionError("Suffixless direct URL override should fail before probing")

    monkeypatch.setattr(refresh, "_probe_stream_url", fail_if_called)

    try:
        refresh.probe_dataset_source(
            dataset,
            timeout=5.0,
            file_overrides={},
            url_overrides={"ONSUD": "https://downloads.example.test/onsud/download"},
        )
    except ValueError as exc:
        assert "Direct URL source has no ingestible file suffix" in str(exc)
    else:
        raise AssertionError("Expected suffixless direct URL override to be rejected")


def test_open_rows_and_rows_from_bytes_support_jsonl(tmp_path: Path) -> None:
    rows_path = tmp_path / "sample.jsonl"
    rows_path.write_text(
        '{"UPRN":"100023336959","LAD24CD":"E08000026"}\n'
        '{"UPRN":"100023336960","LAD24CD":"E08000026"}\n',
        encoding="utf-8",
    )

    with refresh._open_rows(rows_path) as (rows, fieldnames):
        payload = list(rows)

    assert fieldnames == ["UPRN", "LAD24CD"]
    assert payload[0]["UPRN"] == "100023336959"


def test_open_rows_collects_jsonl_fieldnames_from_multiple_rows(tmp_path: Path) -> None:
    rows_path = tmp_path / "sample.jsonl"
    rows_path.write_text(
        '{"UPRN":"100023336959"}\n'
        '{"UPRN":"100023336960","LAD24CD":"E08000026","WARD24CD":"E05001111"}\n',
        encoding="utf-8",
    )

    with refresh._open_rows(rows_path) as (rows, fieldnames):
        payload = list(rows)

    assert fieldnames == ["UPRN", "LAD24CD", "WARD24CD"]
    assert payload[1]["WARD24CD"] == "E05001111"


def test_open_rows_supports_jsonl_inside_zip(tmp_path: Path) -> None:
    archive_path = tmp_path / "sample.zip"
    with refresh.zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr(
            "rows.jsonl",
            '{"UPRN":"100023336959","LAD24CD":"E08000026"}\n',
        )

    with refresh._open_rows(archive_path) as (rows, fieldnames):
        payload = list(rows)

    assert fieldnames == ["UPRN", "LAD24CD"]
    assert payload[0]["LAD24CD"] == "E08000026"


def test_open_rows_chooses_zip_member_with_matching_schema(tmp_path: Path) -> None:
    archive_path = tmp_path / "sample.zip"
    with refresh.zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("README.csv", "note,description\nx,y\n")
        archive.writestr(
            "rows.csv",
            "GEOGRAPHY_CODE,STATUS\nE09000033,current\n",
        )

    dataset = refresh.DatasetConfig(
        dataset_id="RGC",
        dataset_kind="support",
        title="RGC",
        key_type=None,
        derivation_mode=None,
        priority=20,
        release="2025-12",
        resolver=refresh.ResolverConfig("static_file", "", "", "", "", [], [], [], ""),
        required_fields=["code", "name"],
        optional_fields=["status", "code_family", "level"],
        aliases={"name": ["GEOGRAPHY_NAME"], "code": ["GEOGRAPHY_CODE"]},
        defaults={"status": "current"},
    )

    with refresh._open_rows(archive_path, dataset=dataset) as (rows, fieldnames):
        payload = list(rows)

    assert fieldnames == ["GEOGRAPHY_CODE", "STATUS"]
    assert payload[0]["GEOGRAPHY_CODE"] == "E09000033"


def test_select_candidate_url_ignores_page_anchors_and_assets() -> None:
    chosen = refresh._select_candidate_url(
        candidates=[
            "https://example.test/dataset/onsud-latest#main-content",
            "https://hubcdn.arcgis.com/assets/vendor.css",
            "https://open-geography.example/datasets/onsud-latest",
            "https://example.test/downloads/onsud-latest.zip",
        ],
        link_patterns=["onsud", "zip"],
        preferred_suffixes=[".zip", ".csv"],
    )
    assert chosen == "https://example.test/downloads/onsud-latest.zip"


def test_code_reference_normalization_maps_retired_codes_to_current() -> None:
    store = refresh.CodeReferenceStore()
    chd_dataset = refresh.DatasetConfig(
        dataset_id="CHD",
        dataset_kind="support",
        title="CHD",
        key_type=None,
        derivation_mode=None,
        priority=10,
        release="2025-12",
        resolver=refresh.ResolverConfig("static_file", "", "", "", "", [], [], [], ""),
        required_fields=["code", "status"],
        optional_fields=["name", "successor_code", "successor_name", "code_family", "level"],
        aliases={},
        defaults={"status": "retired"},
    )
    rgc_dataset = refresh.DatasetConfig(
        dataset_id="RGC",
        dataset_kind="support",
        title="RGC",
        key_type=None,
        derivation_mode=None,
        priority=20,
        release="2025-12",
        resolver=refresh.ResolverConfig("static_file", "", "", "", "", [], [], [], ""),
        required_fields=["code", "name"],
        optional_fields=["status", "code_family", "level"],
        aliases={},
        defaults={"status": "current"},
    )

    chd_mapping = {
        "code": "GEOGRAPHY_CODE",
        "name": "GEOGRAPHY_NAME",
        "status": "STATUS",
        "successor_code": "SUCCESSOR_CODE",
        "successor_name": "SUCCESSOR_NAME",
        "code_family": "CODE_FAMILY",
        "level": "LEVEL",
    }
    rgc_mapping = {
        "code": "GEOGRAPHY_CODE",
        "name": "GEOGRAPHY_NAME",
        "status": "STATUS",
        "code_family": "CODE_FAMILY",
        "level": "LEVEL",
    }

    retired = refresh._normalize_code_reference_row(
        {
            "GEOGRAPHY_CODE": "E09000044",
            "GEOGRAPHY_NAME": "Westminster (old)",
            "STATUS": "retired",
            "SUCCESSOR_CODE": "E09000033",
            "SUCCESSOR_NAME": "Westminster",
            "CODE_FAMILY": "lad",
            "LEVEL": "local_authority_district",
        },
        dataset=chd_dataset,
        mapping=chd_mapping,
    )
    current = refresh._normalize_code_reference_row(
        {
            "GEOGRAPHY_CODE": "E09000033",
            "GEOGRAPHY_NAME": "Westminster",
            "STATUS": "current",
            "CODE_FAMILY": "lad",
            "LEVEL": "local_authority_district",
        },
        dataset=rgc_dataset,
        mapping=rgc_mapping,
    )
    assert retired is not None and current is not None
    store.add(retired)
    store.add(current)

    normalized = refresh._build_normalized_row(
        {"LAD24CD": "E09000044", "LAD24NM": "Westminster (old)"},
        mapping={"lad_code": "LAD24CD", "lad_name": "LAD24NM"},
        code_references=store,
    )
    assert normalized["geographies"]["lad"]["status"] == "retired"
    assert normalized["geographies"]["lad"]["currentCode"] == "E09000033"
    assert normalized["geographies"]["lad"]["currentName"] == "Westminster"


def test_code_reference_annotation_flags_family_mismatch() -> None:
    store = refresh.CodeReferenceStore()
    record = refresh.CodeReferenceRecord(
        dataset_id="RGC",
        code="E05013806",
        code_family="ward",
        name="St James's",
        status="current",
        successor_code=None,
        successor_name=None,
        level="ward",
        record={},
    )
    store.add(record)
    normalized = refresh._build_normalized_row(
        {"LAD24CD": "E05013806"},
        mapping={"lad_code": "LAD24CD"},
        code_references=store,
    )
    assert normalized["geographies"]["lad"]["status"] == "family_mismatch"
