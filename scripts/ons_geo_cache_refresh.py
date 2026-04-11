#!/usr/bin/env python3
"""Refresh the local ONS geography cache used by ons_geo.* tools."""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import html
import io
import json
import re
import sqlite3
import time
import zipfile
from collections.abc import Iterable, Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urljoin, urlparse

import requests
from openpyxl import load_workbook

from server.ons_geo_cache import (
    KEY_TYPES,
    ensure_schema,
    normalize_derivation_mode,
    normalize_key,
    normalize_postcode,
    normalize_uprn,
)
from server.ons_geo_freshness import (
    load_addressbase_epoch_schedule,
    parse_epoch_from_text,
    summarize_uprn_dataset_freshness,
)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCES_PATH = ROOT / "resources" / "ons_geo_sources.json"
DEFAULT_CACHE_DIR = ROOT / "data" / "cache" / "ons_geo"
DEFAULT_INDEX_PATH = ROOT / "resources" / "ons_geo_cache_index.json"
DEFAULT_DB_NAME = "ons_geo_cache.sqlite"

_MONTH_NAME_RE = re.compile(
    r"(?P<month>January|February|March|April|May|June|July|August|September|October|"
    r"November|December)\s+(?P<year>20\d{2})",
    re.IGNORECASE,
)
_EPOCH_RE = re.compile(r"Epoch[\s_-]+(?P<epoch>\d+)", re.IGNORECASE)
_HREF_RE = re.compile(r"""href=["'](?P<href>[^"']+)["']""", re.IGNORECASE)
_IGNORED_CANDIDATE_SUFFIXES = {
    ".css",
    ".gif",
    ".ico",
    ".jpeg",
    ".jpg",
    ".js",
    ".jsonld",
    ".png",
    ".svg",
    ".webp",
    ".woff",
    ".woff2",
}
_SUPPORTED_RESOLVERS = {
    "hosted_table_arcgis",
    "portal_release_file",
    "static_file",
    "direct_url",
}
_DIRECT_INGEST_SUFFIXES = {".zip", ".csv", ".json", ".ndjson", ".jsonl", ".gz"}
_JSON_LINES_SCHEMA_SAMPLE_ROWS = 50
_CONTENT_DISPOSITION_FILENAME_RE = re.compile(
    r"""filename\*?=(?:UTF-8''|")?(?P<name>[^";]+)""",
    re.IGNORECASE,
)
_SQLITE_REFRESH_INSERT_BATCH = 50_000
_RGC_XLSX_FIELDNAMES = [
    "GEOGRAPHY_CODE",
    "GEOGRAPHY_NAME",
    "STATUS",
    "CODE_FAMILY",
    "LEVEL",
]
_RGC_SHEET_SKIP_NAMES = {"RGC", "Metadata_for_geography_listings", "For_Scotland"}
_RGC_SHEET_ANNOTATIONS = {
    "OA": ("oa", "output_area"),
    "LSOA": ("lsoa", "lower_layer_super_output_area"),
    "DZ": ("lsoa", "lower_layer_super_output_area"),
    "MSOA": ("msoa", "middle_layer_super_output_area"),
    "SDZ": ("msoa", "middle_layer_super_output_area"),
    "WD": ("ward", "ward"),
    "CMWD": ("ward", "ward"),
    "DEA": ("ward", "ward"),
    "UA": ("lad", "local_authority_district"),
    "NMD": ("lad", "local_authority_district"),
    "MD": ("lad", "local_authority_district"),
    "LONB": ("lad", "local_authority_district"),
    "CTY": ("lad", "local_authority_district"),
    "MCTY": ("lad", "local_authority_district"),
    "LGD": ("lad", "local_authority_district"),
    "LAD": ("lad", "local_authority_district"),
    "GLAD": ("lad", "local_authority_district"),
    "GLTLA": ("lad", "local_authority_district"),
    "CMLAD": ("lad", "local_authority_district"),
    "CMLAD21": ("lad", "local_authority_district"),
    "CTRY": ("country", "country"),
    "RGN": ("region", "region"),
}

_SEMANTIC_FIELD_REGEX: dict[str, re.Pattern[str]] = {
    "postcode": re.compile(r"^(pcds|pcd|postcode|post_code)$", re.IGNORECASE),
    "uprn": re.compile(r"^uprn$", re.IGNORECASE),
    "oa_code": re.compile(r"^(oa\d{2}cd|oacd)$", re.IGNORECASE),
    "oa_name": re.compile(r"^(oa\d{2}nm|oanm)$", re.IGNORECASE),
    "lsoa_code": re.compile(r"^(lsoa\d{2}cd|lsoacd)$", re.IGNORECASE),
    "lsoa_name": re.compile(r"^(lsoa\d{2}nm|lsoanm)$", re.IGNORECASE),
    "msoa_code": re.compile(r"^(msoa\d{2}cd|msoacd)$", re.IGNORECASE),
    "msoa_name": re.compile(r"^(msoa\d{2}nm|msoanm)$", re.IGNORECASE),
    "lad_code": re.compile(r"^(lad\d{2}cd|ladcd)$", re.IGNORECASE),
    "lad_name": re.compile(r"^(lad\d{2}nm|ladnm)$", re.IGNORECASE),
    "ward_code": re.compile(r"^((wd|ward)\d{2}cd|wdcd|wardcd)$", re.IGNORECASE),
    "ward_name": re.compile(r"^((wd|ward)\d{2}nm|wdnm|wardnm)$", re.IGNORECASE),
    "country_code": re.compile(r"^((ctry|country)\d{2}cd|ctrycd|countrycd)$", re.IGNORECASE),
    "country_name": re.compile(r"^((ctry|country)\d{2}nm|ctrynm|countrynm)$", re.IGNORECASE),
    "region_code": re.compile(r"^((rgn|region)\d{2}cd|rgncd|regioncd)$", re.IGNORECASE),
    "region_name": re.compile(r"^((rgn|region)\d{2}nm|rgnnm|regionnm)$", re.IGNORECASE),
    "postal_delivery": re.compile(
        r"^(postal_delivery|postaldelivery|delivery_point|receives_post|postal_address)$",
        re.IGNORECASE,
    ),
}

_SEMANTIC_ALIAS_PATTERNS: dict[str, tuple[re.Pattern[str], ...]] = {
    "lad_code": (re.compile(r"local authority district code", re.IGNORECASE),),
    "lad_name": (re.compile(r"local authority district name", re.IGNORECASE),),
    "ward_code": (re.compile(r"ward code", re.IGNORECASE),),
    "ward_name": (re.compile(r"ward name", re.IGNORECASE),),
    "country_code": (re.compile(r"country code", re.IGNORECASE),),
    "country_name": (re.compile(r"country name", re.IGNORECASE),),
    "region_code": (re.compile(r"region code", re.IGNORECASE),),
    "region_name": (re.compile(r"region name", re.IGNORECASE),),
    "oa_code": (re.compile(r"output area code", re.IGNORECASE),),
    "oa_name": (re.compile(r"output area name", re.IGNORECASE),),
    "lsoa_code": (re.compile(r"lower layer super output area code", re.IGNORECASE),),
    "lsoa_name": (re.compile(r"lower layer super output area name", re.IGNORECASE),),
    "msoa_code": (re.compile(r"middle layer super output area code", re.IGNORECASE),),
    "msoa_name": (re.compile(r"middle layer super output area name", re.IGNORECASE),),
    "postcode": (re.compile(r"postcode", re.IGNORECASE),),
    "uprn": (re.compile(r"\buprn\b", re.IGNORECASE),),
    "postal_delivery": (re.compile(r"delivery", re.IGNORECASE),),
}

_GEOGRAPHY_GROUPS: dict[str, tuple[str, str]] = {
    "oa": ("oa_code", "oa_name"),
    "lsoa": ("lsoa_code", "lsoa_name"),
    "msoa": ("msoa_code", "msoa_name"),
    "lad": ("lad_code", "lad_name"),
    "ward": ("ward_code", "ward_name"),
    "country": ("country_code", "country_name"),
    "region": ("region_code", "region_name"),
}

_MAIN_PRODUCT_SEMANTICS = {
    "postcode",
    "uprn",
    "postal_delivery",
    *{item for pair in _GEOGRAPHY_GROUPS.values() for item in pair},
}


@dataclass(frozen=True)
class ResolverConfig:
    resolver_type: str
    landing_url: str
    metadata_url: str
    query_url: str
    static_path: str
    preferred_suffixes: list[str]
    link_patterns: list[str]
    release_patterns: list[str]
    discovery_api_url: str


@dataclass(frozen=True)
class DatasetConfig:
    dataset_id: str
    dataset_kind: str
    title: str
    key_type: str | None
    derivation_mode: str | None
    priority: int
    release: str
    resolver: ResolverConfig
    required_fields: list[str]
    optional_fields: list[str]
    aliases: dict[str, list[str]]
    defaults: dict[str, Any]


@dataclass(frozen=True)
class ResolvedSource:
    dataset_id: str
    source_path: Path
    source_format: str
    resolved_source_url: str | None
    resolved_release: str | None
    retrieved_at: str
    metadata: dict[str, Any]
    schema_fields: list[str]
    field_aliases: dict[str, str]


@dataclass(frozen=True)
class SourceProbe:
    dataset_id: str
    resolver_type: str
    source_format: str | None
    resolved_source_url: str | None
    resolved_release: str | None
    retrieved_at: str
    metadata: dict[str, Any]
    schema_fields: list[str]
    field_aliases: dict[str, str]
    schema_probe_status: str
    warning: str | None


@dataclass(frozen=True)
class CodeReferenceRecord:
    dataset_id: str
    code: str
    code_family: str | None
    name: str | None
    status: str
    successor_code: str | None
    successor_name: str | None
    level: str | None
    record: dict[str, Any]


class CodeReferenceStore:
    def __init__(self) -> None:
        self._current_by_code: dict[str, CodeReferenceRecord] = {}
        self._history_by_code: dict[str, CodeReferenceRecord] = {}

    def add(self, record: CodeReferenceRecord) -> None:
        status = record.status
        if status == "current":
            existing = self._current_by_code.get(record.code)
            if existing is None or existing.dataset_id != "RGC":
                self._current_by_code[record.code] = record
            return
        self._history_by_code[record.code] = record

    def annotate(self, code: str, expected_family: str | None) -> dict[str, Any]:
        record = self._current_by_code.get(code)
        if record is not None:
            matches = _family_matches(expected_family, record.code_family)
            return {
                "status": "current" if matches or record.code_family is None else "family_mismatch",
                "sourceDataset": record.dataset_id,
                "codeFamily": record.code_family,
                "currentCode": record.code,
                "currentName": record.name,
                "successorCode": None,
                "successorName": None,
                "level": record.level,
            }

        history = self._history_by_code.get(code)
        if history is None:
            return {
                "status": "unknown",
                "sourceDataset": None,
                "codeFamily": None,
                "currentCode": None,
                "currentName": None,
                "successorCode": None,
                "successorName": None,
                "level": None,
            }

        matches = _family_matches(expected_family, history.code_family)
        successor = self._current_by_code.get(history.successor_code or "")
        return {
            "status": (
                "retired"
                if matches or history.code_family is None
                else "family_mismatch"
            ),
            "sourceDataset": history.dataset_id,
            "codeFamily": history.code_family,
            "currentCode": successor.code if successor else history.successor_code,
            "currentName": successor.name if successor else history.successor_name,
            "successorCode": history.successor_code,
            "successorName": history.successor_name,
            "level": history.level,
        }


def _family_matches(expected_family: str | None, actual_family: str | None) -> bool:
    if not expected_family or not actual_family:
        return True
    return expected_family.strip().lower() == actual_family.strip().lower()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _display_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return str(resolved)


def _parse_map_args(values: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for item in values:
        key, sep, raw_value = item.partition("=")
        if not sep:
            raise ValueError(f"Expected KEY=VALUE format, got: {item}")
        key = key.strip().upper()
        value = raw_value.strip()
        if not key or not value:
            raise ValueError(f"Expected KEY=VALUE format, got: {item}")
        mapping[key] = value
    return mapping


def _load_dataset_config(raw: dict[str, Any], *, dataset_kind: str) -> DatasetConfig | None:
    dataset_id = str(raw.get("id") or "").strip().upper()
    title = str(raw.get("title") or dataset_id).strip()
    if not dataset_id or not title:
        return None

    key_type: str | None = None
    derivation_mode: str | None = None
    if dataset_kind == "product":
        raw_key_type = str(raw.get("keyType") or "").strip().lower()
        key_type = raw_key_type if raw_key_type in KEY_TYPES else None
        derivation_mode = normalize_derivation_mode(str(raw.get("derivationMode") or ""))
        if key_type is None or derivation_mode is None:
            return None

    priority_raw = raw.get("priority", 100)
    priority = int(priority_raw) if isinstance(priority_raw, int) else 100
    release = str(raw.get("release") or "unknown")

    resolver_raw = raw.get("resolver")
    source_raw = raw.get("source")
    if not isinstance(resolver_raw, dict):
        resolver_raw = {}
        if isinstance(source_raw, dict):
            download_url = str(source_raw.get("downloadUrl") or "").strip()
            if download_url:
                resolver_raw = {"type": "direct_url", "downloadUrl": download_url}
    resolver_type = str(resolver_raw.get("type") or "").strip().lower()
    if resolver_type not in _SUPPORTED_RESOLVERS:
        return None

    resolver = ResolverConfig(
        resolver_type=resolver_type,
        landing_url=str(
            resolver_raw.get("landingUrl")
            or resolver_raw.get("downloadUrl")
            or ""
        ).strip(),
        metadata_url=str(resolver_raw.get("metadataUrl") or "").strip(),
        query_url=str(resolver_raw.get("queryUrl") or "").strip(),
        static_path=str(resolver_raw.get("path") or "").strip(),
        preferred_suffixes=[
            str(item).strip().lower()
            for item in resolver_raw.get("preferredSuffixes", [])
            if str(item).strip()
        ]
        if isinstance(resolver_raw.get("preferredSuffixes"), list)
        else [],
        link_patterns=[
            str(item).strip()
            for item in resolver_raw.get("linkPatterns", [])
            if str(item).strip()
        ]
        if isinstance(resolver_raw.get("linkPatterns"), list)
        else [],
        release_patterns=[
            str(item).strip()
            for item in resolver_raw.get("releasePatterns", [])
            if str(item).strip()
        ]
        if isinstance(resolver_raw.get("releasePatterns"), list)
        else [],
        discovery_api_url=str(resolver_raw.get("discoveryApiUrl") or "").strip(),
    )

    semantic_raw = raw.get("semanticFields")
    aliases: dict[str, list[str]] = {}
    required_fields: list[str] = []
    optional_fields: list[str] = []
    defaults: dict[str, Any] = {}
    if isinstance(semantic_raw, dict):
        aliases_raw = semantic_raw.get("aliases")
        if isinstance(aliases_raw, dict):
            for key, values in aliases_raw.items():
                if not isinstance(values, list):
                    continue
                aliases[str(key).strip()] = [
                    str(item).strip() for item in values if str(item).strip()
                ]
        required_fields = [
            str(item).strip()
            for item in semantic_raw.get("required", [])
            if str(item).strip()
        ]
        optional_fields = [
            str(item).strip()
            for item in semantic_raw.get("optional", [])
            if str(item).strip()
        ]
        defaults_value = semantic_raw.get("defaults")
        defaults = defaults_value if isinstance(defaults_value, dict) else {}

    if not aliases and isinstance(raw.get("fieldCandidates"), dict):
        legacy_fields = raw["fieldCandidates"]
        if isinstance(legacy_fields.get("key"), list):
            legacy_key = "postcode" if key_type == "postcode" else "uprn"
            aliases[legacy_key] = [
                str(item).strip() for item in legacy_fields["key"] if str(item).strip()
            ]

    if dataset_kind == "product" and key_type and key_type not in aliases:
        aliases[key_type] = ["pcds", "pcd", "postcode"] if key_type == "postcode" else ["uprn"]
    if dataset_kind == "product" and key_type and key_type not in required_fields:
        required_fields = [key_type, *required_fields]

    return DatasetConfig(
        dataset_id=dataset_id,
        dataset_kind=dataset_kind,
        title=title,
        key_type=key_type,
        derivation_mode=derivation_mode,
        priority=priority,
        release=release,
        resolver=resolver,
        required_fields=required_fields,
        optional_fields=optional_fields,
        aliases=aliases,
        defaults=defaults,
    )


def load_manifest(path: Path) -> tuple[str, list[DatasetConfig], list[DatasetConfig]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("sources manifest must be an object")
    version = str(payload.get("version") or "unknown")

    product_items = payload.get("products")
    support_items = payload.get("supportProducts")
    if not isinstance(product_items, list):
        raise ValueError("sources manifest must include a products list")
    if support_items is None:
        support_items = []
    if not isinstance(support_items, list):
        raise ValueError("supportProducts must be a list when present")

    products = [
        cfg
        for cfg in (
            _load_dataset_config(item, dataset_kind="product")
            for item in product_items
            if isinstance(item, dict)
        )
        if cfg is not None
    ]
    support_products = [
        cfg
        for cfg in (
            _load_dataset_config(item, dataset_kind="support")
            for item in support_items
            if isinstance(item, dict)
        )
        if cfg is not None
    ]
    return version, products, support_products


def _safe_release_fragment(value: str | None, fallback: str) -> str:
    raw = str(value or fallback).strip() or fallback
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", raw)
    safe = safe.strip("_")
    return safe or fallback


def _format_detected_release(
    *,
    month_match: re.Match[str] | None,
    epoch_match: re.Match[str] | None,
) -> str | None:
    if month_match and epoch_match:
        return (
            f"{month_match.group('month').title()} {month_match.group('year')} "
            f"(Epoch {epoch_match.group('epoch')})"
        )
    if epoch_match:
        return f"Epoch {epoch_match.group('epoch')}"
    if month_match:
        return f"{month_match.group('month').title()} {month_match.group('year')}"
    return None


def _detect_release(text: str, explicit_patterns: list[str]) -> str | None:
    normalized_text = text.replace("-", " ").replace("_", " ")
    overall_month_match = _MONTH_NAME_RE.search(text) or _MONTH_NAME_RE.search(normalized_text)
    overall_epoch_match = _EPOCH_RE.search(text) or _EPOCH_RE.search(normalized_text)
    for pattern in explicit_patterns:
        try:
            match = re.search(pattern, text, re.IGNORECASE) or re.search(
                pattern, normalized_text, re.IGNORECASE
            )
        except re.error:
            continue
        if match:
            if match.groupdict():
                joined = " ".join(
                    value
                    for value in match.groupdict().values()
                    if isinstance(value, str) and value
                )
                candidate = joined.strip() or match.group(0).strip()
            else:
                candidate = match.group(0).strip()
            candidate_month_match = _MONTH_NAME_RE.search(candidate)
            candidate_epoch_match = _EPOCH_RE.search(candidate)
            if (candidate_month_match or candidate_epoch_match) and (
                overall_month_match or overall_epoch_match
            ):
                combined = _format_detected_release(
                    month_match=overall_month_match or candidate_month_match,
                    epoch_match=overall_epoch_match or candidate_epoch_match,
                )
                if combined:
                    return combined
            return candidate
    return _format_detected_release(
        month_match=overall_month_match,
        epoch_match=overall_epoch_match,
    )


def _response_text(resp: requests.Response) -> str:
    return resp.text if isinstance(resp.text, str) else resp.content.decode("utf-8", "replace")


def _download(url: str, destination: Path, timeout: float) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with requests.get(url, timeout=timeout, stream=True) as resp:
        resp.raise_for_status()
        with destination.open("wb") as handle:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    handle.write(chunk)
    return destination


def _collect_json_urls(value: Any, out: list[str]) -> None:
    if isinstance(value, dict):
        for key, item in value.items():
            lower_key = str(key).lower()
            if lower_key.endswith("url") and isinstance(item, str) and item.strip():
                out.append(item.strip())
            _collect_json_urls(item, out)
        return
    if isinstance(value, list):
        for item in value:
            _collect_json_urls(item, out)


def _arcgis_item_data_url(value: str | None) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    parsed = urlparse(raw)
    if parsed.scheme not in {"http", "https"}:
        return None
    if "arcgis.com" not in parsed.netloc.lower():
        return None

    item_id: str | None = None
    if parsed.path.lower().endswith("/home/item.html"):
        item_id = parse_qs(parsed.query).get("id", [None])[0]
    elif "/sharing/rest/content/items/" in parsed.path.lower():
        parts = [part for part in parsed.path.split("/") if part]
        try:
            index = [part.lower() for part in parts].index("items")
        except ValueError:
            index = -1
        if index >= 0 and index + 1 < len(parts):
            item_id = parts[index + 1]

    normalized_id = str(item_id or "").strip()
    if not normalized_id:
        return None
    return f"https://www.arcgis.com/sharing/rest/content/items/{normalized_id}/data"


def _collect_special_discovery_urls(value: Any, out: list[str]) -> None:
    if isinstance(value, dict):
        extra_key = str(value.get("key") or "").strip().lower()
        extra_value = value.get("value")
        if extra_key == "guid" and isinstance(extra_value, str):
            data_url = _arcgis_item_data_url(extra_value)
            if data_url:
                out.append(data_url)
        for key, item in value.items():
            lower_key = str(key).lower()
            if lower_key == "guid" and isinstance(item, str):
                data_url = _arcgis_item_data_url(item)
                if data_url:
                    out.append(data_url)
            _collect_special_discovery_urls(item, out)
        return
    if isinstance(value, list):
        for item in value:
            _collect_special_discovery_urls(item, out)


def _parse_ckan_datetime(value: str | None) -> tuple[float, str]:
    raw = str(value or "").strip()
    if not raw:
        return (0.0, "")
    normalized = raw.replace("Z", "+00:00")
    try:
        return (datetime.fromisoformat(normalized).timestamp(), raw)
    except ValueError:
        return (0.0, raw)


def _data_gov_dataset_url(package_name: str | None) -> str | None:
    name = str(package_name or "").strip().strip("/")
    if not name:
        return None
    return f"https://www.data.gov.uk/dataset/{name}"


def _is_auxiliary_package(package: dict[str, Any]) -> bool:
    combined = " ".join(
        str(package.get(field) or "").strip().lower()
        for field in ("title", "name")
    )
    return any(
        token in combined
        for token in (
            "user guide",
            "guidance",
            "methodology",
            "metadata",
            "technical guide",
        )
    )


def _score_ckan_package_result(
    dataset: DatasetConfig,
    package: dict[str, Any],
) -> tuple[int, int, float, str]:
    title = str(package.get("title") or "").strip()
    name = str(package.get("name") or "").strip()
    combined = " ".join(part for part in (title, name) if part)
    pattern_score = sum(
        1
        for pattern in dataset.resolver.link_patterns
        if re.search(pattern, combined, re.IGNORECASE)
    )
    if _is_auxiliary_package(package):
        pattern_score -= 100
    epoch_score = parse_epoch_from_text(title, name) or 0
    modified_score, modified_raw = _parse_ckan_datetime(
        str(package.get("metadata_modified") or package.get("metadata_created") or "")
    )
    return (pattern_score, epoch_score, modified_score, modified_raw or combined)


def _extract_ckan_discovery_context(
    dataset: DatasetConfig,
    api_payload: dict[str, Any],
) -> tuple[list[str], str | None, str | None, dict[str, Any]]:
    metadata: dict[str, Any] = {}
    discovery_urls: list[str] = []
    release_hint: str | None = None
    landing_override: str | None = None

    result = api_payload.get("result")
    selected_package: dict[str, Any] | None = None
    if isinstance(result, dict) and isinstance(result.get("results"), list):
        candidates = [item for item in result["results"] if isinstance(item, dict)]
        if candidates:
            selected_package = sorted(
                candidates,
                key=lambda item: _score_ckan_package_result(dataset, item),
                reverse=True,
            )[0]
            metadata["selectedPackage"] = {
                "name": selected_package.get("name"),
                "title": selected_package.get("title"),
                "metadataModified": selected_package.get("metadata_modified"),
                "metadataCreated": selected_package.get("metadata_created"),
            }
            _collect_json_urls(selected_package, discovery_urls)
            _collect_special_discovery_urls(selected_package, discovery_urls)
            landing_override = _data_gov_dataset_url(selected_package.get("name"))
            release_hint = _detect_release(
                json.dumps(selected_package, ensure_ascii=True),
                dataset.resolver.release_patterns,
            )
    elif isinstance(result, dict):
        metadata["selectedPackage"] = {
            "name": result.get("name"),
            "title": result.get("title"),
            "metadataModified": result.get("metadata_modified"),
            "metadataCreated": result.get("metadata_created"),
        }
        _collect_json_urls(result, discovery_urls)
        _collect_special_discovery_urls(result, discovery_urls)
        landing_override = _data_gov_dataset_url(result.get("name"))
        release_hint = _detect_release(
            json.dumps(result, ensure_ascii=True),
            dataset.resolver.release_patterns,
        )
    else:
        _collect_json_urls(api_payload, discovery_urls)
        _collect_special_discovery_urls(api_payload, discovery_urls)
        release_hint = _detect_release(
            json.dumps(api_payload, ensure_ascii=True),
            dataset.resolver.release_patterns,
        )

    return discovery_urls, release_hint, landing_override, metadata


def _extract_links_from_html(base_url: str, body: str) -> list[str]:
    links: list[str] = []
    for match in _HREF_RE.finditer(body):
        href = html.unescape(match.group("href")).strip()
        if not href:
            continue
        links.append(urljoin(base_url, href))
    return links


def _normalize_candidate_url(url: str) -> str | None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return None
    suffix = Path(parsed.path).suffix.lower()
    if suffix in _IGNORED_CANDIDATE_SUFFIXES:
        return None
    if parsed.fragment and not suffix and not parsed.query:
        return None
    return parsed._replace(fragment="").geturl()


def _select_candidate_url(
    *,
    candidates: list[str],
    link_patterns: list[str],
    preferred_suffixes: list[str],
) -> str | None:
    normalized_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        normalized = _normalize_candidate_url(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        normalized_candidates.append(normalized)

    if not normalized_candidates:
        return None

    def _score(url: str) -> tuple[int, int, int, str]:
        normalized = url.lower()
        parsed = urlparse(url)
        pattern_score = sum(
            1 for pattern in link_patterns if re.search(pattern, url, re.IGNORECASE)
        )
        suffix_score = 0
        for index, suffix in enumerate(preferred_suffixes):
            if normalized.endswith(suffix.lower()):
                suffix_score = len(preferred_suffixes) - index
                break
        downloadish_score = 1 if any(
            token in parsed.path.lower()
            for token in ("/download", "/downloads/", "/resource/", "/datasets/")
        ) else 0
        return (pattern_score, suffix_score, downloadish_score, url)

    filtered = [
        item
        for item in normalized_candidates
        if (
            not link_patterns
            or any(re.search(pattern, item, re.IGNORECASE) for pattern in link_patterns)
        )
    ]
    pool = filtered or normalized_candidates
    return sorted(pool, key=_score, reverse=True)[0]


def _resolve_portal_release_file(
    dataset: DatasetConfig,
    *,
    raw_root: Path,
    timeout: float,
) -> ResolvedSource:
    landing_url = dataset.resolver.landing_url
    if not landing_url:
        raise ValueError("portal_release_file resolver requires landingUrl")

    metadata: dict[str, Any] = {"landingUrl": landing_url}
    discovery_urls: list[str] = []
    release_hint: str | None = None
    landing_url_to_fetch = landing_url

    if dataset.resolver.discovery_api_url:
        with requests.get(dataset.resolver.discovery_api_url, timeout=timeout) as resp:
            resp.raise_for_status()
            api_payload = resp.json()
        metadata["discoveryApiUrl"] = dataset.resolver.discovery_api_url
        metadata["discoveryApiPayload"] = api_payload
        (
            discovery_urls,
            release_hint,
            landing_override,
            discovery_metadata,
        ) = _extract_ckan_discovery_context(dataset, api_payload)
        metadata.update(discovery_metadata)
        if landing_override:
            landing_url_to_fetch = landing_override
            metadata["resolvedLandingUrl"] = landing_override

    chosen_url, chosen_suffix = _select_direct_discovery_url(
        dataset=dataset,
        discovery_urls=discovery_urls,
        timeout=timeout,
    )
    body = ""
    if chosen_url is None or chosen_suffix is None:
        with requests.get(landing_url_to_fetch, timeout=timeout) as resp:
            resp.raise_for_status()
            body = _response_text(resp)
            metadata["landingContentType"] = resp.headers.get("content-type")
            metadata["landingFetchedUrl"] = landing_url_to_fetch

        html_links = _extract_links_from_html(landing_url_to_fetch, body)
        chosen_url = _select_candidate_url(
            candidates=[*discovery_urls, *html_links],
            link_patterns=dataset.resolver.link_patterns,
            preferred_suffixes=dataset.resolver.preferred_suffixes,
        )
        if chosen_url is None:
            parsed = urlparse(landing_url_to_fetch)
            if Path(parsed.path).suffix:
                chosen_url = landing_url_to_fetch
            else:
                raise ValueError(
                    "Could not resolve a release-file download URL from landing metadata"
                )

        candidate_suffix = Path(urlparse(chosen_url).path).suffix.lower()
        preferred_suffixes = set(dataset.resolver.preferred_suffixes)
        direct_download_suffixes = _DIRECT_INGEST_SUFFIXES
        if (
            candidate_suffix not in preferred_suffixes
            and candidate_suffix not in direct_download_suffixes
        ):
            with requests.get(chosen_url, timeout=timeout) as resp:
                resp.raise_for_status()
                chosen_body = _response_text(resp)
            secondary_links = _extract_links_from_html(chosen_url, chosen_body)
            refined_url = _select_candidate_url(
                candidates=secondary_links,
                link_patterns=dataset.resolver.link_patterns,
                preferred_suffixes=dataset.resolver.preferred_suffixes,
            )
            if refined_url:
                chosen_url = refined_url

        chosen_suffix = _validate_direct_ingest_suffix(chosen_url, timeout=timeout)

    release = release_hint or _detect_release(
        body + "\n" + chosen_url,
        dataset.resolver.release_patterns,
    )
    release = release or dataset.release
    suffix = chosen_suffix
    raw_dir = raw_root / dataset.dataset_id / _safe_release_fragment(release, "latest")
    filename = Path(urlparse(chosen_url).path).name
    if not filename or Path(filename).suffix.lower() != suffix.lower():
        filename = f"{dataset.dataset_id.lower()}{suffix}"
    local_path = raw_dir / filename
    _download(chosen_url, local_path, timeout=timeout)
    return ResolvedSource(
        dataset_id=dataset.dataset_id,
        source_path=local_path,
        source_format=chosen_suffix.lstrip(".").lower(),
        resolved_source_url=chosen_url,
        resolved_release=release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata=metadata,
        schema_fields=[],
        field_aliases={},
    )


def _probe_stream_url(url: str, *, timeout: float) -> None:
    with requests.get(url, timeout=timeout, stream=True) as resp:
        resp.raise_for_status()


def _arcgis_page_signature(
    rows: list[dict[str, Any]],
    *,
    object_id_field: str,
) -> str:
    if object_id_field:
        object_ids = [
            row.get(object_id_field)
            for row in rows
            if row.get(object_id_field) is not None
        ]
        if object_ids:
            return json.dumps(object_ids, ensure_ascii=True, separators=(",", ":"))
    return json.dumps(rows, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def _validate_direct_ingest_suffix(
    chosen_url: str,
    *,
    label: str = "Resolved release asset",
    timeout: float | None = None,
) -> str:
    chosen_suffix = Path(urlparse(chosen_url).path).suffix.lower()
    if not chosen_suffix and timeout is not None:
        with requests.get(chosen_url, timeout=timeout, stream=True) as resp:
            resp.raise_for_status()
            content_disposition = str(resp.headers.get("content-disposition") or "").strip()
            match = _CONTENT_DISPOSITION_FILENAME_RE.search(content_disposition)
            if match:
                chosen_suffix = Path(match.group("name").strip()).suffix.lower()
            if not chosen_suffix:
                content_type = str(resp.headers.get("content-type") or "").split(";", 1)[0].strip()
                content_type_map = {
                    "application/csv": ".csv",
                    "application/gzip": ".gz",
                    "application/json": ".json",
                    "application/ndjson": ".ndjson",
                    "application/x-ndjson": ".ndjson",
                    "application/zip": ".zip",
                    "application/x-zip-compressed": ".zip",
                    "text/csv": ".csv",
                }
                chosen_suffix = content_type_map.get(content_type.lower(), "")
    if not chosen_suffix:
        raise ValueError(
            f"{label} has no ingestible file suffix; "
            "supported direct formats are zip/csv/json/ndjson/jsonl/gz."
        )
    if chosen_suffix not in _DIRECT_INGEST_SUFFIXES:
        raise ValueError(
            f"{label} uses unsupported format {chosen_suffix}; "
            "supported direct formats are zip/csv/json/ndjson/jsonl/gz."
        )
    return chosen_suffix


def _select_direct_discovery_url(
    *,
    dataset: DatasetConfig,
    discovery_urls: list[str],
    timeout: float,
) -> tuple[str | None, str | None]:
    valid_direct_urls: list[tuple[str, str]] = []
    for candidate in discovery_urls:
        normalized = _normalize_candidate_url(candidate)
        if not normalized:
            continue
        try:
            direct_suffix = _validate_direct_ingest_suffix(normalized, timeout=timeout)
        except (ValueError, requests.RequestException):
            continue
        valid_direct_urls.append((normalized, direct_suffix))

    if not valid_direct_urls:
        return None, None

    direct_url = _select_candidate_url(
        candidates=[item[0] for item in valid_direct_urls],
        link_patterns=dataset.resolver.link_patterns,
        preferred_suffixes=dataset.resolver.preferred_suffixes,
    )
    if direct_url is None:
        return None, None
    direct_suffix = next(
        (suffix for url, suffix in valid_direct_urls if url == direct_url),
        None,
    )
    if direct_suffix is None:
        return None, None
    return direct_url, direct_suffix


def _probe_portal_release_file(
    dataset: DatasetConfig,
    *,
    timeout: float,
) -> SourceProbe:
    landing_url = dataset.resolver.landing_url
    if not landing_url:
        raise ValueError("portal_release_file resolver requires landingUrl")

    metadata: dict[str, Any] = {"landingUrl": landing_url}
    discovery_urls: list[str] = []
    release_hint: str | None = None
    landing_url_to_fetch = landing_url

    if dataset.resolver.discovery_api_url:
        with requests.get(dataset.resolver.discovery_api_url, timeout=timeout) as resp:
            resp.raise_for_status()
            api_payload = resp.json()
        metadata["discoveryApiUrl"] = dataset.resolver.discovery_api_url
        metadata["discoveryApiPayload"] = api_payload
        (
            discovery_urls,
            release_hint,
            landing_override,
            discovery_metadata,
        ) = _extract_ckan_discovery_context(dataset, api_payload)
        metadata.update(discovery_metadata)
        if landing_override:
            landing_url_to_fetch = landing_override
            metadata["resolvedLandingUrl"] = landing_override

    chosen_url, chosen_suffix = _select_direct_discovery_url(
        dataset=dataset,
        discovery_urls=discovery_urls,
        timeout=timeout,
    )
    body = ""
    if chosen_url is None or chosen_suffix is None:
        with requests.get(landing_url_to_fetch, timeout=timeout) as resp:
            resp.raise_for_status()
            body = _response_text(resp)
            metadata["landingContentType"] = resp.headers.get("content-type")
            metadata["landingFetchedUrl"] = landing_url_to_fetch

        html_links = _extract_links_from_html(landing_url_to_fetch, body)
        chosen_url = _select_candidate_url(
            candidates=[*discovery_urls, *html_links],
            link_patterns=dataset.resolver.link_patterns,
            preferred_suffixes=dataset.resolver.preferred_suffixes,
        )
        if chosen_url is None:
            parsed = urlparse(landing_url_to_fetch)
            if Path(parsed.path).suffix:
                chosen_url = landing_url_to_fetch
            else:
                raise ValueError(
                    "Could not resolve a release-file download URL from landing metadata"
                )

        candidate_suffix = Path(urlparse(chosen_url).path).suffix.lower()
        preferred_suffixes = set(dataset.resolver.preferred_suffixes)
        direct_download_suffixes = _DIRECT_INGEST_SUFFIXES
        if (
            candidate_suffix not in preferred_suffixes
            and candidate_suffix not in direct_download_suffixes
        ):
            with requests.get(chosen_url, timeout=timeout) as resp:
                resp.raise_for_status()
                chosen_body = _response_text(resp)
            secondary_links = _extract_links_from_html(chosen_url, chosen_body)
            refined_url = _select_candidate_url(
                candidates=secondary_links,
                link_patterns=dataset.resolver.link_patterns,
                preferred_suffixes=dataset.resolver.preferred_suffixes,
            )
            if refined_url:
                chosen_url = refined_url

        chosen_suffix = _validate_direct_ingest_suffix(chosen_url, timeout=timeout)

    _probe_stream_url(chosen_url, timeout=timeout)
    release = release_hint or _detect_release(
        body + "\n" + chosen_url,
        dataset.resolver.release_patterns,
    )
    release = release or dataset.release
    source_format = chosen_suffix.lstrip(".").lower()
    return SourceProbe(
        dataset_id=dataset.dataset_id,
        resolver_type=dataset.resolver.resolver_type,
        source_format=source_format,
        resolved_source_url=chosen_url,
        resolved_release=release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata=metadata,
        schema_fields=[],
        field_aliases={},
        schema_probe_status="unavailable_remote_archive",
        warning=(
            "Remote archive schema was not inspected during live validation; "
            "run a real refresh to validate fields."
        ),
    )


def _resolve_static_file(dataset: DatasetConfig) -> ResolvedSource:
    path_text = dataset.resolver.static_path
    if not path_text:
        raise ValueError("static_file resolver requires path")
    source_path = Path(path_text).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"Static source path does not exist: {source_path}")
    suffix = source_path.suffix.lstrip(".").lower() or "csv"
    return ResolvedSource(
        dataset_id=dataset.dataset_id,
        source_path=source_path,
        source_format=suffix,
        resolved_source_url=None,
        resolved_release=dataset.release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata={},
        schema_fields=[],
        field_aliases={},
    )


def _resolve_direct_url(
    dataset: DatasetConfig,
    *,
    raw_root: Path,
    timeout: float,
) -> ResolvedSource:
    url = dataset.resolver.landing_url
    if not url:
        raise ValueError("direct_url resolver requires downloadUrl/landingUrl")
    release = dataset.release
    suffix = _validate_direct_ingest_suffix(url, label="Direct URL source", timeout=timeout)
    raw_dir = raw_root / dataset.dataset_id / _safe_release_fragment(release, "latest")
    filename = Path(urlparse(url).path).name
    if not filename or Path(filename).suffix.lower() != suffix.lower():
        filename = f"{dataset.dataset_id.lower()}{suffix}"
    local_path = raw_dir / filename
    _download(url, local_path, timeout=timeout)
    return ResolvedSource(
        dataset_id=dataset.dataset_id,
        source_path=local_path,
        source_format=suffix.lstrip(".").lower(),
        resolved_source_url=url,
        resolved_release=release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata={},
        schema_fields=[],
        field_aliases={},
    )


def _probe_direct_url(
    dataset: DatasetConfig,
    *,
    timeout: float,
) -> SourceProbe:
    url = dataset.resolver.landing_url
    if not url:
        raise ValueError("direct_url resolver requires downloadUrl/landingUrl")
    suffix = _validate_direct_ingest_suffix(url, label="Direct URL source", timeout=timeout)
    _probe_stream_url(url, timeout=timeout)
    return SourceProbe(
        dataset_id=dataset.dataset_id,
        resolver_type=dataset.resolver.resolver_type,
        source_format=suffix.lstrip(".").lower(),
        resolved_source_url=url,
        resolved_release=dataset.release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata={},
        schema_fields=[],
        field_aliases={},
        schema_probe_status="reachable_only",
        warning="Direct URL source was reached, but no schema probe was attempted.",
    )


def _arcgis_metadata_url(dataset: DatasetConfig) -> str:
    if dataset.resolver.metadata_url:
        return dataset.resolver.metadata_url
    if dataset.resolver.query_url:
        return dataset.resolver.query_url.split("/query", 1)[0] + "?f=json"
    if dataset.resolver.landing_url:
        return dataset.resolver.landing_url.rstrip("/") + "?f=json"
    raise ValueError("hosted_table_arcgis resolver requires metadataUrl or queryUrl")


def _arcgis_query_url(dataset: DatasetConfig, metadata_url: str) -> str:
    if dataset.resolver.query_url:
        return dataset.resolver.query_url
    if metadata_url.endswith("?f=json"):
        return metadata_url[: -len("?f=json")].rstrip("/") + "/query"
    return metadata_url.rstrip("/") + "/query"


def _resolve_hosted_table_arcgis(
    dataset: DatasetConfig,
    *,
    raw_root: Path,
    timeout: float,
) -> ResolvedSource:
    metadata_url = _arcgis_metadata_url(dataset)
    query_url = _arcgis_query_url(dataset, metadata_url)
    with requests.get(metadata_url, timeout=timeout) as resp:
        resp.raise_for_status()
        metadata = resp.json()

    fields = metadata.get("fields", [])
    field_aliases: dict[str, str] = {}
    schema_fields: list[str] = []
    if isinstance(fields, list):
        for field in fields:
            if not isinstance(field, dict):
                continue
            name = str(field.get("name") or "").strip()
            if not name:
                continue
            schema_fields.append(name)
            alias = str(field.get("alias") or "").strip()
            if alias:
                field_aliases[name] = alias

    release = _detect_release(
        json.dumps(metadata, ensure_ascii=True),
        dataset.resolver.release_patterns,
    ) or _detect_release(str(metadata.get("name") or ""), dataset.resolver.release_patterns)
    release = release or dataset.release
    raw_dir = raw_root / dataset.dataset_id / _safe_release_fragment(release, "latest")
    raw_dir.mkdir(parents=True, exist_ok=True)
    metadata_path = raw_dir / "metadata.json"
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=True, indent=2) + "\n",
        encoding="utf-8",
    )
    rows_path = raw_dir / "rows.ndjson"

    max_record_count = metadata.get("maxRecordCount")
    if isinstance(max_record_count, int) and max_record_count > 0:
        page_size = min(max_record_count, 2000)
    else:
        page_size = 1000

    offset = 0
    object_id_field = str(metadata.get("objectIdField") or "").strip()
    last_object_id: Any | None = None
    seen_page_signatures: set[str] = set()
    with rows_path.open("w", encoding="utf-8") as handle:
        while True:
            params: dict[str, str | int] = {
                "outFields": "*",
                "f": "json",
                "returnGeometry": "false",
                "resultRecordCount": page_size,
            }
            if object_id_field:
                if last_object_id is None:
                    params["where"] = "1=1"
                else:
                    params["where"] = f"{object_id_field} > {last_object_id}"
                params["orderByFields"] = object_id_field
            else:
                params["where"] = "1=1"
                params["resultOffset"] = offset
            with requests.get(query_url, timeout=timeout, params=params) as resp:
                resp.raise_for_status()
                page = resp.json()
            features = page.get("features", [])
            if not isinstance(features, list):
                raise ValueError("ArcGIS query response did not contain a features list")
            page_rows: list[dict[str, Any]] = []
            written = 0
            for feature in features:
                if not isinstance(feature, dict):
                    continue
                attributes = feature.get("attributes")
                if not isinstance(attributes, dict):
                    continue
                page_rows.append(attributes)
            if page_rows:
                page_signature = _arcgis_page_signature(
                    page_rows,
                    object_id_field=object_id_field,
                )
                if page_signature in seen_page_signatures:
                    raise ValueError(
                        "ArcGIS query pagination made no progress; "
                        "endpoint may be ignoring resultOffset."
                    )
                seen_page_signatures.add(page_signature)
                if object_id_field:
                    next_object_id = page_rows[-1].get(object_id_field)
                    if next_object_id is None:
                        raise ValueError(
                            f"ArcGIS page did not include {object_id_field} values "
                            "required for keyset pagination."
                        )
                    if last_object_id is not None and next_object_id == last_object_id:
                        raise ValueError(
                            "ArcGIS query pagination made no progress; "
                            "endpoint may be ignoring objectId keyset pagination."
                        )
                    last_object_id = next_object_id
            for attributes in page_rows:
                handle.write(
                    json.dumps(attributes, ensure_ascii=True, separators=(",", ":")) + "\n"
                )
                written += 1
            if written == 0:
                break
            offset += written
            if written < page_size and not page.get("exceededTransferLimit"):
                break

    return ResolvedSource(
        dataset_id=dataset.dataset_id,
        source_path=rows_path,
        source_format="ndjson",
        resolved_source_url=query_url,
        resolved_release=release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata=metadata,
        schema_fields=schema_fields,
        field_aliases=field_aliases,
    )


def _probe_hosted_table_arcgis(
    dataset: DatasetConfig,
    *,
    timeout: float,
) -> SourceProbe:
    metadata_url = _arcgis_metadata_url(dataset)
    query_url = _arcgis_query_url(dataset, metadata_url)
    with requests.get(metadata_url, timeout=timeout) as resp:
        resp.raise_for_status()
        metadata = resp.json()

    fields = metadata.get("fields", [])
    field_aliases: dict[str, str] = {}
    schema_fields: list[str] = []
    if isinstance(fields, list):
        for field in fields:
            if not isinstance(field, dict):
                continue
            name = str(field.get("name") or "").strip()
            if not name:
                continue
            schema_fields.append(name)
            alias = str(field.get("alias") or "").strip()
            if alias:
                field_aliases[name] = alias

    release = _detect_release(
        json.dumps(metadata, ensure_ascii=True),
        dataset.resolver.release_patterns,
    ) or _detect_release(str(metadata.get("name") or ""), dataset.resolver.release_patterns)
    release = release or dataset.release
    return SourceProbe(
        dataset_id=dataset.dataset_id,
        resolver_type=dataset.resolver.resolver_type,
        source_format="arcgis_feature_service",
        resolved_source_url=query_url,
        resolved_release=release,
        retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        metadata=metadata,
        schema_fields=schema_fields,
        field_aliases=field_aliases,
        schema_probe_status="metadata",
        warning=None,
    )


def resolve_dataset_source(
    dataset: DatasetConfig,
    *,
    raw_root: Path,
    timeout: float,
    file_overrides: dict[str, str],
    url_overrides: dict[str, str],
) -> ResolvedSource:
    override_path = file_overrides.get(dataset.dataset_id)
    if override_path:
        source_path = Path(override_path).expanduser().resolve()
        if not source_path.exists():
            raise FileNotFoundError(f"Missing override file: {source_path}")
        suffix = source_path.suffix.lstrip(".").lower() or "csv"
        return ResolvedSource(
            dataset_id=dataset.dataset_id,
            source_path=source_path,
            source_format=suffix,
            resolved_source_url=None,
            resolved_release=dataset.release,
            retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            metadata={"override": "file"},
            schema_fields=[],
            field_aliases={},
        )

    override_url = url_overrides.get(dataset.dataset_id)
    if override_url:
        override_dataset = DatasetConfig(
            dataset_id=dataset.dataset_id,
            dataset_kind=dataset.dataset_kind,
            title=dataset.title,
            key_type=dataset.key_type,
            derivation_mode=dataset.derivation_mode,
            priority=dataset.priority,
            release=dataset.release,
            resolver=ResolverConfig(
                resolver_type="direct_url",
                landing_url=override_url,
                metadata_url="",
                query_url="",
                static_path="",
                preferred_suffixes=[],
                link_patterns=[],
                release_patterns=dataset.resolver.release_patterns,
                discovery_api_url="",
            ),
            required_fields=dataset.required_fields,
            optional_fields=dataset.optional_fields,
            aliases=dataset.aliases,
            defaults=dataset.defaults,
        )
        return _resolve_direct_url(override_dataset, raw_root=raw_root, timeout=timeout)

    resolver_type = dataset.resolver.resolver_type
    if resolver_type == "hosted_table_arcgis":
        return _resolve_hosted_table_arcgis(dataset, raw_root=raw_root, timeout=timeout)
    if resolver_type == "portal_release_file":
        return _resolve_portal_release_file(dataset, raw_root=raw_root, timeout=timeout)
    if resolver_type == "static_file":
        return _resolve_static_file(dataset)
    if resolver_type == "direct_url":
        return _resolve_direct_url(dataset, raw_root=raw_root, timeout=timeout)
    raise ValueError(f"Unsupported resolver type: {resolver_type}")


def probe_dataset_source(
    dataset: DatasetConfig,
    *,
    timeout: float,
    file_overrides: dict[str, str],
    url_overrides: dict[str, str],
) -> SourceProbe:
    override_path = file_overrides.get(dataset.dataset_id)
    if override_path:
        source_path = Path(override_path).expanduser().resolve()
        if not source_path.exists():
            raise FileNotFoundError(f"Missing override file: {source_path}")
        suffix = source_path.suffix.lstrip(".").lower() or "csv"
        return SourceProbe(
            dataset_id=dataset.dataset_id,
            resolver_type="static_file",
            source_format=suffix,
            resolved_source_url=None,
            resolved_release=dataset.release,
            retrieved_at=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            metadata={"override": "file", "sourcePath": str(source_path)},
            schema_fields=[],
            field_aliases={},
            schema_probe_status="local_file",
            warning=None,
        )

    override_url = url_overrides.get(dataset.dataset_id)
    if override_url:
        override_dataset = DatasetConfig(
            dataset_id=dataset.dataset_id,
            dataset_kind=dataset.dataset_kind,
            title=dataset.title,
            key_type=dataset.key_type,
            derivation_mode=dataset.derivation_mode,
            priority=dataset.priority,
            release=dataset.release,
            resolver=ResolverConfig(
                resolver_type="direct_url",
                landing_url=override_url,
                metadata_url="",
                query_url="",
                static_path="",
                preferred_suffixes=[],
                link_patterns=[],
                release_patterns=dataset.resolver.release_patterns,
                discovery_api_url="",
            ),
            required_fields=dataset.required_fields,
            optional_fields=dataset.optional_fields,
            aliases=dataset.aliases,
            defaults=dataset.defaults,
        )
        return _probe_direct_url(override_dataset, timeout=timeout)

    resolver_type = dataset.resolver.resolver_type
    if resolver_type == "hosted_table_arcgis":
        return _probe_hosted_table_arcgis(dataset, timeout=timeout)
    if resolver_type == "portal_release_file":
        return _probe_portal_release_file(dataset, timeout=timeout)
    if resolver_type == "static_file":
        static = _resolve_static_file(dataset)
        return SourceProbe(
            dataset_id=dataset.dataset_id,
            resolver_type="static_file",
            source_format=static.source_format,
            resolved_source_url=static.resolved_source_url,
            resolved_release=static.resolved_release,
            retrieved_at=static.retrieved_at,
            metadata={"sourcePath": str(static.source_path)},
            schema_fields=[],
            field_aliases={},
            schema_probe_status="local_file",
            warning=None,
        )
    if resolver_type == "direct_url":
        return _probe_direct_url(dataset, timeout=timeout)
    raise ValueError(f"Unsupported resolver type: {resolver_type}")


def _rgc_sheet_annotation(sheet_name: str) -> tuple[str | None, str | None]:
    suffix = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
    return _RGC_SHEET_ANNOTATIONS.get(suffix, (None, None))


@contextmanager
def _open_rows(
    path: Path,
    *,
    dataset: DatasetConfig | None = None,
    metadata_aliases: dict[str, str] | None = None,
) -> Iterator[tuple[Iterator[dict[str, Any]], list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            members = sorted(
                name
                for name in archive.namelist()
                if not name.endswith("/")
                and (
                    name.lower().endswith(".csv")
                    or name.lower().endswith(".json")
                    or name.lower().endswith(".ndjson")
                    or name.lower().endswith(".jsonl")
                    or (
                        dataset is not None
                        and dataset.dataset_id == "RGC"
                        and name.lower().endswith(".xlsx")
                    )
                )
            )
            if not members:
                raise ValueError(
                    f"No CSV, JSON, or supported XLSX file found in zip archive: {path}"
                )
            chosen_member = _select_archive_member(
                archive=archive,
                members=members,
                dataset=dataset,
                metadata_aliases=metadata_aliases or {},
            )
            with archive.open(chosen_member, "r") as raw:
                if chosen_member.lower().endswith(".xlsx"):
                    with _rows_from_xlsx_stream(
                        raw,
                        name=chosen_member,
                        dataset=dataset,
                    ) as payload:
                        yield payload
                else:
                    with _rows_from_binary_stream(raw, name=chosen_member) as payload:
                        yield payload
        return
    if path.suffix.lower() == ".xlsx":
        with path.open("rb") as stream:
            with _rows_from_xlsx_stream(stream, name=path.name, dataset=dataset) as payload:
                yield payload
        return
    if suffix == ".gz":
        with gzip.open(path, "rb") as stream:
            with _rows_from_binary_stream(stream, name=path.stem) as payload:
                yield payload
        return
    with path.open("rb") as stream:
        with _rows_from_binary_stream(stream, name=path.name) as payload:
            yield payload


@contextmanager
def _rows_from_xlsx_stream(
    stream: Any,
    *,
    name: str,
    dataset: DatasetConfig | None,
) -> Iterator[tuple[Iterator[dict[str, Any]], list[str]]]:
    if dataset is None or dataset.dataset_id != "RGC":
        raise ValueError(
            f"Unsupported XLSX source for {name}; only RGC workbook archives are supported."
        )
    workbook = load_workbook(io.BytesIO(stream.read()), read_only=True, data_only=True)
    try:
        def _iter_rows() -> Iterator[dict[str, Any]]:
            for sheet_name in workbook.sheetnames:
                if sheet_name in _RGC_SHEET_SKIP_NAMES:
                    continue
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)
                try:
                    header_row = next(rows)
                except StopIteration:
                    continue
                headers = {
                    str(value).strip().upper(): index
                    for index, value in enumerate(header_row)
                    if value is not None and str(value).strip()
                }
                code_index = headers.get("GEOGCD")
                name_index = headers.get("GEOGNM")
                status_index = headers.get("STATUS")
                if code_index is None or name_index is None:
                    continue
                code_family, level = _rgc_sheet_annotation(sheet_name)
                for row in rows:
                    if not isinstance(row, tuple):
                        continue
                    code_value = row[code_index] if code_index < len(row) else None
                    name_value = row[name_index] if name_index < len(row) else None
                    status_value = (
                        row[status_index]
                        if status_index is not None and status_index < len(row)
                        else None
                    )
                    code = str(code_value or "").strip().upper()
                    geography_name = str(name_value or "").strip()
                    if not code or not geography_name:
                        continue
                    payload = {
                        "GEOGRAPHY_CODE": code,
                        "GEOGRAPHY_NAME": geography_name,
                        "STATUS": str(status_value).strip() if status_value is not None else None,
                        "CODE_FAMILY": code_family,
                        "LEVEL": level,
                    }
                    yield payload

        yield _iter_rows(), list(_RGC_XLSX_FIELDNAMES)
    finally:
        workbook.close()
    return


@contextmanager
def _rows_from_binary_stream(
    stream: Any,
    *,
    name: str,
) -> Iterator[tuple[Iterator[dict[str, Any]], list[str]]]:
    lower_name = name.lower()
    if lower_name.endswith(".json"):
        text_stream = io.TextIOWrapper(stream, encoding="utf-8-sig")
        try:
            payload = json.load(text_stream)
        finally:
            text_stream.detach()
        rows, fieldnames = _rows_from_json_payload(payload)
        yield iter(rows), fieldnames
        return
    if lower_name.endswith(".ndjson") or lower_name.endswith(".jsonl"):
        text_stream = io.TextIOWrapper(stream, encoding="utf-8-sig")

        def _iter_json_lines() -> Iterator[dict[str, Any]]:
            for line in text_stream:
                if not line.strip():
                    continue
                payload = json.loads(line)
                if isinstance(payload, dict):
                    yield payload

        try:
            iterator = _iter_json_lines()
            buffered_rows: list[dict[str, Any]] = []
            for row in iterator:
                buffered_rows.append(row)
                if len(buffered_rows) >= _JSON_LINES_SCHEMA_SAMPLE_ROWS:
                    break
            if not buffered_rows:
                yield iter(()), []
                return
            fieldnames = _collect_fieldnames(buffered_rows)

            def _iter_rows() -> Iterator[dict[str, Any]]:
                yield from buffered_rows
                yield from iterator

            yield _iter_rows(), fieldnames
        finally:
            text_stream.detach()
        return

    text_stream = io.TextIOWrapper(stream, encoding="utf-8-sig", newline="")
    reader = csv.DictReader(text_stream)
    fieldnames = list(reader.fieldnames or [])
    try:
        yield iter(reader), [str(item) for item in fieldnames if isinstance(item, str)]
    finally:
        text_stream.detach()
    return


@contextmanager
def _rows_from_bytes(
    content: bytes,
    *,
    name: str,
) -> Iterator[tuple[Iterator[dict[str, Any]], list[str]]]:
    with _rows_from_binary_stream(io.BytesIO(content), name=name) as payload:
        yield payload


def _rows_from_json_payload(payload: Any) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        if isinstance(payload.get("features"), list):
            for feature in payload["features"]:
                if not isinstance(feature, dict):
                    continue
                attributes = feature.get("attributes")
                if isinstance(attributes, dict):
                    rows.append(attributes)
        elif isinstance(payload.get("records"), list):
            rows = [item for item in payload["records"] if isinstance(item, dict)]
        elif isinstance(payload.get("rows"), list):
            rows = [item for item in payload["rows"] if isinstance(item, dict)]
        else:
            rows = [payload]
    elif isinstance(payload, list):
        rows = [item for item in payload if isinstance(item, dict)]
    fieldnames = _collect_fieldnames(rows)
    return rows, fieldnames


def _merged_fieldnames(
    inferred_fieldnames: list[str],
    schema_fieldnames: list[str],
) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for name in [*inferred_fieldnames, *schema_fieldnames]:
        if not isinstance(name, str):
            continue
        if name in seen:
            continue
        seen.add(name)
        merged.append(name)
    return merged


def _select_archive_member(
    *,
    archive: zipfile.ZipFile,
    members: list[str],
    dataset: DatasetConfig | None,
    metadata_aliases: dict[str, str],
) -> str:
    if dataset is None or len(members) == 1:
        return members[0]

    scored_members: list[tuple[int, int, int, str]] = []
    for name in members:
        with archive.open(name, "r") as raw:
            if name.lower().endswith(".xlsx"):
                payload_context = _rows_from_xlsx_stream(raw, name=name, dataset=dataset)
            else:
                payload_context = _rows_from_binary_stream(raw, name=name)
            with payload_context as (_rows_iter, fieldnames):
                _mapping, validation = _build_field_mapping(
                    dataset,
                    fieldnames=fieldnames,
                    metadata_aliases=metadata_aliases,
                )
        required_found = len(validation["requiredFound"])
        required_missing = len(validation["requiredMissing"])
        optional_found = len(validation["optionalFound"])
        scored_members.append(
            (required_found, -required_missing, optional_found, name)
        )

    scored_members.sort(reverse=True)
    return scored_members[0][3]


def _collect_fieldnames(rows: Iterable[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        for key in row:
            text = str(key).strip()
            if text and text not in seen:
                seen.append(text)
    return seen


def _row_value_ci(row: dict[str, Any], candidates: list[str]) -> str | None:
    lower_lookup = {
        str(key).strip().lower(): value
        for key, value in row.items()
        if isinstance(key, str)
    }
    for candidate in candidates:
        value = lower_lookup.get(candidate.lower())
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return None


def _metadata_alias_matches(semantic_name: str, field_alias: str) -> bool:
    patterns = _SEMANTIC_ALIAS_PATTERNS.get(semantic_name, ())
    return any(pattern.search(field_alias) for pattern in patterns)


def _match_field(
    fieldnames: list[str],
    *,
    semantic_name: str,
    aliases: dict[str, list[str]],
    metadata_aliases: dict[str, str],
) -> str | None:
    lookup = {name.lower(): name for name in fieldnames}
    explicit_aliases = aliases.get(semantic_name, [])
    for candidate in explicit_aliases:
        chosen = lookup.get(candidate.lower())
        if chosen:
            return chosen

    regex = _SEMANTIC_FIELD_REGEX.get(semantic_name)
    if regex is not None:
        for fieldname in fieldnames:
            if regex.match(fieldname):
                return fieldname

    for fieldname in fieldnames:
        alias = metadata_aliases.get(fieldname, "")
        if alias and _metadata_alias_matches(semantic_name, alias):
            return fieldname
    return None


def _schema_fingerprint(fieldnames: list[str], metadata_aliases: dict[str, str]) -> str:
    payload = [
        {"name": fieldname, "alias": metadata_aliases.get(fieldname)}
        for fieldname in sorted(fieldnames)
    ]
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _build_field_mapping(
    dataset: DatasetConfig,
    *,
    fieldnames: list[str],
    metadata_aliases: dict[str, str],
) -> tuple[dict[str, str], dict[str, Any]]:
    mapping: dict[str, str] = {}
    used_fields: set[str] = set()
    for semantic_name in [*dataset.required_fields, *dataset.optional_fields]:
        matched = _match_field(
            fieldnames,
            semantic_name=semantic_name,
            aliases=dataset.aliases,
            metadata_aliases=metadata_aliases,
        )
        if matched:
            mapping[semantic_name] = matched
            used_fields.add(matched)

    validation: dict[str, Any] = {
        "requiredFound": [name for name in dataset.required_fields if name in mapping],
        "requiredMissing": [name for name in dataset.required_fields if name not in mapping],
        "optionalFound": [name for name in dataset.optional_fields if name in mapping],
        "unknownFields": [name for name in fieldnames if name not in used_fields],
    }
    validation["status"] = "ok" if not validation["requiredMissing"] else "schema_drift"
    return mapping, validation


def _normalize_support_status(value: str | None, dataset: DatasetConfig) -> str:
    if value is None or not value.strip():
        default_status = str(dataset.defaults.get("status") or "").strip().lower()
        return default_status or "unknown"
    normalized = value.strip().lower().replace(" ", "_")
    if normalized in {"live", "active"}:
        return "current"
    if normalized in {"terminated", "closed"}:
        return "retired"
    return normalized


def _normalize_support_family(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().lower().replace(" ", "_")
    return normalized or None


def _normalize_code_reference_row(
    row: dict[str, Any],
    *,
    dataset: DatasetConfig,
    mapping: dict[str, str],
) -> CodeReferenceRecord | None:
    code_field = mapping.get("code")
    if not code_field:
        return None
    code = str(row.get(code_field) or "").strip().upper()
    if not code:
        return None
    name = str(row.get(mapping["name"]) or "").strip() if "name" in mapping else None
    status = _normalize_support_status(
        (
            str(row.get(mapping["status"]))
            if "status" in mapping and row.get(mapping["status"]) is not None
            else None
        ),
        dataset,
    )
    successor_code = (
        str(row.get(mapping["successor_code"]) or "").strip().upper()
        if "successor_code" in mapping
        else None
    ) or None
    successor_name = (
        str(row.get(mapping["successor_name"]) or "").strip()
        if "successor_name" in mapping
        else None
    ) or None
    code_family = _normalize_support_family(

            str(row.get(mapping["code_family"]))
            if "code_family" in mapping and row.get(mapping["code_family"]) is not None
            else None

    )
    level = (
        str(row.get(mapping["level"]) or "").strip()
        if "level" in mapping
        else None
    ) or None
    return CodeReferenceRecord(
        dataset_id=dataset.dataset_id,
        code=code,
        code_family=code_family,
        name=name or None,
        status=status,
        successor_code=successor_code,
        successor_name=successor_name,
        level=level,
        record=row,
    )


def _normalize_postal_delivery(value: str | None) -> int | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in {"1", "true", "t", "yes", "y"}:
        return 1
    if normalized in {"0", "false", "f", "no", "n"}:
        return 0
    return None


def _build_normalized_row(
    row: dict[str, Any],
    *,
    mapping: dict[str, str],
    code_references: CodeReferenceStore,
) -> dict[str, Any]:
    semantic_fields: dict[str, Any] = {}
    for semantic_name, raw_field in mapping.items():
        value = row.get(raw_field)
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        if semantic_name == "postcode":
            normalized = normalize_postcode(text)
            semantic_fields[semantic_name] = normalized or text
        elif semantic_name == "uprn":
            normalized = normalize_uprn(text)
            semantic_fields[semantic_name] = normalized or text
        else:
            semantic_fields[semantic_name] = text

    geographies: dict[str, dict[str, Any]] = {}
    status_summary: dict[str, int] = {}
    for family, (code_key, name_key) in _GEOGRAPHY_GROUPS.items():
        code = semantic_fields.get(code_key)
        name = semantic_fields.get(name_key)
        if not code and not name:
            continue
        annotation = code_references.annotate(str(code), family) if code else {
            "status": "unknown",
            "sourceDataset": None,
            "codeFamily": family,
            "currentCode": None,
            "currentName": None,
            "successorCode": None,
            "successorName": None,
            "level": None,
        }
        status = str(annotation.get("status") or "unknown")
        status_summary[status] = status_summary.get(status, 0) + 1
        geographies[family] = {
            "code": code,
            "name": name,
            "currentCode": annotation.get("currentCode") or code,
            "currentName": annotation.get("currentName") or name,
            "status": status,
            "sourceDataset": annotation.get("sourceDataset"),
            "codeFamily": annotation.get("codeFamily"),
            "successorCode": annotation.get("successorCode"),
            "successorName": annotation.get("successorName"),
            "level": annotation.get("level"),
        }

    return {
        "semanticFields": semantic_fields,
        "geographies": geographies,
        "codeStatusSummary": status_summary,
    }


def _insert_uprn_index_row(
    *,
    conn: sqlite3.Connection,
    dataset: DatasetConfig,
    normalized_row: dict[str, Any],
    cached_at: str,
) -> None:
    row_values = _uprn_index_row_values(
        dataset=dataset,
        normalized_row=normalized_row,
        cached_at=cached_at,
    )
    if row_values is None:
        return
    conn.execute(
        """
        INSERT OR REPLACE INTO ons_geo_uprn_index (
            product_id,
            derivation_mode,
            uprn,
            postcode,
            oa_code,
            lsoa_code,
            msoa_code,
            lad_code,
            lad_name,
            ward_code,
            ward_name,
            country_code,
            country_name,
            region_code,
            region_name,
            postal_delivery,
            geographies_json,
            cached_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        row_values,
    )


def _uprn_index_row_values(
    *,
    dataset: DatasetConfig,
    normalized_row: dict[str, Any],
    cached_at: str,
) -> tuple[Any, ...] | None:
    semantic = normalized_row.get("semanticFields", {})
    if not isinstance(semantic, dict):
        return
    uprn = semantic.get("uprn")
    if not isinstance(uprn, str) or not uprn:
        return
    geographies = normalized_row.get("geographies", {})
    if not isinstance(geographies, dict):
        geographies = {}

    def _geo_value(family: str, key: str) -> str | None:
        entry = geographies.get(family)
        if not isinstance(entry, dict):
            return None
        value = entry.get(key)
        return str(value).strip() if isinstance(value, str) and value.strip() else None

    postal_delivery = semantic.get("postal_delivery")
    postal_delivery_value = None
    if isinstance(postal_delivery, str):
        postal_delivery_value = _normalize_postal_delivery(postal_delivery)
    elif isinstance(postal_delivery, int):
        postal_delivery_value = postal_delivery

    return (
        dataset.dataset_id,
        dataset.derivation_mode,
        uprn,
        semantic.get("postcode"),
        _geo_value("oa", "currentCode"),
        _geo_value("lsoa", "currentCode"),
        _geo_value("msoa", "currentCode"),
        _geo_value("lad", "currentCode"),
        _geo_value("lad", "currentName"),
        _geo_value("ward", "currentCode"),
        _geo_value("ward", "currentName"),
        _geo_value("country", "currentCode"),
        _geo_value("country", "currentName"),
        _geo_value("region", "currentCode"),
        _geo_value("region", "currentName"),
        postal_delivery_value,
        json.dumps(geographies, ensure_ascii=True, separators=(",", ":")),
        cached_at,
    )


def _flush_pending_rows(
    conn: sqlite3.Connection,
    statement: str,
    pending: list[tuple[Any, ...]],
) -> None:
    if not pending:
        return
    conn.executemany(statement, pending)
    pending.clear()


def _ingest_support_dataset(
    *,
    conn: sqlite3.Connection,
    dataset: DatasetConfig,
    resolved: ResolvedSource,
    max_rows: int | None,
    code_references: CodeReferenceStore,
) -> tuple[int, dict[str, Any]]:
    now_iso = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    inserted = 0
    insert_sql = """
        INSERT OR REPLACE INTO ons_geo_code_reference (
            dataset_id,
            code,
            code_family,
            name,
            status,
            successor_code,
            successor_name,
            level,
            record_json,
            ingested_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    pending_records: list[tuple[Any, ...]] = []
    with _open_rows(
        resolved.source_path,
        dataset=dataset,
        metadata_aliases=resolved.field_aliases,
    ) as (rows_iter, fieldnames):
        metadata_aliases = resolved.field_aliases
        available_fieldnames = _merged_fieldnames(fieldnames, resolved.schema_fields)
        mapping, validation = _build_field_mapping(
            dataset,
            fieldnames=available_fieldnames,
            metadata_aliases=metadata_aliases,
        )
        if validation["requiredMissing"]:
            raise ValueError(
                "Support dataset schema drift: missing required fields "
                f"{validation['requiredMissing']}"
            )
        conn.execute(
            "DELETE FROM ons_geo_code_reference WHERE dataset_id = ?",
            (dataset.dataset_id,),
        )
        for row in rows_iter:
            if not isinstance(row, dict):
                continue
            record = _normalize_code_reference_row(row, dataset=dataset, mapping=mapping)
            if record is None:
                continue
            code_references.add(record)
            pending_records.append(
                (
                    record.dataset_id,
                    record.code,
                    record.code_family,
                    record.name,
                    record.status,
                    record.successor_code,
                    record.successor_name,
                    record.level,
                    json.dumps(record.record, ensure_ascii=True, separators=(",", ":")),
                    now_iso,
                )
            )
            inserted += 1
            if len(pending_records) >= _SQLITE_REFRESH_INSERT_BATCH:
                _flush_pending_rows(conn, insert_sql, pending_records)
            if max_rows is not None and inserted >= max_rows:
                break
        _flush_pending_rows(conn, insert_sql, pending_records)

    schema_validation = {
        **validation,
        "schemaFingerprint": _schema_fingerprint(
            available_fieldnames,
            resolved.field_aliases,
        ),
    }
    _upsert_product_metadata(
        conn=conn,
        dataset=dataset,
        resolved=resolved,
        record_count=inserted,
        schema_validation=schema_validation,
        status="ingested",
    )
    return inserted, schema_validation


def _ingest_main_dataset(
    *,
    conn: sqlite3.Connection,
    dataset: DatasetConfig,
    resolved: ResolvedSource,
    max_rows: int | None,
    code_references: CodeReferenceStore,
) -> tuple[int, dict[str, Any], str | None]:
    now_iso = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    inserted = 0
    chosen_key_field: str | None = None
    rows_insert_sql = """
        INSERT OR REPLACE INTO ons_geo_rows (
            product_id,
            key_type,
            key_norm,
            derivation_mode,
            release,
            source_name,
            product_priority,
            row_json,
            normalized_json,
            cached_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    uprn_index_insert_sql = """
        INSERT OR REPLACE INTO ons_geo_uprn_index (
            product_id,
            derivation_mode,
            uprn,
            postcode,
            oa_code,
            lsoa_code,
            msoa_code,
            lad_code,
            lad_name,
            ward_code,
            ward_name,
            country_code,
            country_name,
            region_code,
            region_name,
            postal_delivery,
            geographies_json,
            cached_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    pending_rows: list[tuple[Any, ...]] = []
    pending_uprn_rows: list[tuple[Any, ...]] = []
    with _open_rows(
        resolved.source_path,
        dataset=dataset,
        metadata_aliases=resolved.field_aliases,
    ) as (rows_iter, fieldnames):
        metadata_aliases = resolved.field_aliases
        available_fieldnames = _merged_fieldnames(fieldnames, resolved.schema_fields)
        mapping, validation = _build_field_mapping(
            dataset,
            fieldnames=available_fieldnames,
            metadata_aliases=metadata_aliases,
        )
        if validation["requiredMissing"]:
            raise ValueError(
                "Main dataset schema drift: missing required fields "
                f"{validation['requiredMissing']}"
            )
        chosen_key_field = mapping.get(dataset.key_type or "")
        if not chosen_key_field:
            raise ValueError(f"Could not resolve key field for {dataset.key_type}")

        conn.execute("DELETE FROM ons_geo_rows WHERE product_id = ?", (dataset.dataset_id,))
        if dataset.key_type == "uprn":
            conn.execute(
                "DELETE FROM ons_geo_uprn_index WHERE product_id = ?",
                (dataset.dataset_id,),
            )

        for row in rows_iter:
            if not isinstance(row, dict):
                continue
            raw_key = row.get(chosen_key_field)
            if raw_key is None:
                continue
            key_norm = normalize_key(dataset.key_type or "", str(raw_key))
            if key_norm is None:
                continue
            normalized_row = _build_normalized_row(
                row,
                mapping=mapping,
                code_references=code_references,
            )
            pending_rows.append(
                (
                    dataset.dataset_id,
                    dataset.key_type,
                    key_norm,
                    dataset.derivation_mode,
                    resolved.resolved_release or dataset.release,
                    dataset.title,
                    dataset.priority,
                    json.dumps(row, ensure_ascii=True, separators=(",", ":")),
                    json.dumps(normalized_row, ensure_ascii=True, separators=(",", ":")),
                    now_iso,
                )
            )
            if dataset.key_type == "uprn":
                uprn_index_row = _uprn_index_row_values(
                    dataset=dataset,
                    normalized_row=normalized_row,
                    cached_at=now_iso,
                )
                if uprn_index_row is not None:
                    pending_uprn_rows.append(uprn_index_row)
            inserted += 1
            if len(pending_rows) >= _SQLITE_REFRESH_INSERT_BATCH:
                _flush_pending_rows(conn, rows_insert_sql, pending_rows)
                _flush_pending_rows(conn, uprn_index_insert_sql, pending_uprn_rows)
            if max_rows is not None and inserted >= max_rows:
                break
        _flush_pending_rows(conn, rows_insert_sql, pending_rows)
        _flush_pending_rows(conn, uprn_index_insert_sql, pending_uprn_rows)

    schema_validation = {
        **validation,
        "schemaFingerprint": _schema_fingerprint(
            available_fieldnames,
            resolved.field_aliases,
        ),
    }
    _upsert_product_metadata(
        conn=conn,
        dataset=dataset,
        resolved=resolved,
        record_count=inserted,
        schema_validation=schema_validation,
        status="ingested",
    )
    return inserted, schema_validation, chosen_key_field


def _upsert_product_metadata(
    *,
    conn: sqlite3.Connection,
    dataset: DatasetConfig,
    resolved: ResolvedSource,
    record_count: int,
    schema_validation: dict[str, Any],
    status: str,
) -> None:
    now_iso = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    conn.execute(
        """
        INSERT OR REPLACE INTO ons_geo_products (
            product_id,
            dataset_kind,
            key_type,
            derivation_mode,
            release,
            resolved_release,
            source_name,
            source_path,
            resolved_source_url,
            resolver_type,
            source_format,
            source_sha256,
            schema_fingerprint,
            schema_validation_json,
            record_count,
            status,
            ingested_at,
            retrieved_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            dataset.dataset_id,
            dataset.dataset_kind,
            dataset.key_type,
            dataset.derivation_mode,
            dataset.release,
            resolved.resolved_release or dataset.release,
            dataset.title,
            str(resolved.source_path),
            resolved.resolved_source_url,
            dataset.resolver.resolver_type,
            resolved.source_format,
            _sha256_file(resolved.source_path),
            schema_validation.get("schemaFingerprint"),
            json.dumps(schema_validation, ensure_ascii=True, separators=(",", ":")),
            record_count,
            status,
            now_iso,
            resolved.retrieved_at,
        ),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Refresh local ONS geography cache and supporting code-history datasets."
    )
    parser.add_argument(
        "--sources",
        default=str(DEFAULT_SOURCES_PATH),
        help="Path to sources manifest JSON",
    )
    parser.add_argument(
        "--cache-dir",
        default=str(DEFAULT_CACHE_DIR),
        help="Directory for cache artifacts",
    )
    parser.add_argument(
        "--index-path",
        default=str(DEFAULT_INDEX_PATH),
        help="Path to write cache index JSON",
    )
    parser.add_argument("--db-name", default=DEFAULT_DB_NAME, help="SQLite database filename")
    parser.add_argument(
        "--product-file",
        action="append",
        default=[],
        help="Dataset input override as DATASET_ID=/path/to/file (repeatable)",
    )
    parser.add_argument(
        "--product-url",
        action="append",
        default=[],
        help="Dataset URL override as DATASET_ID=https://... (repeatable)",
    )
    parser.add_argument("--max-rows", type=int, default=None, help="Optional row limit per dataset")
    parser.add_argument("--timeout", type=float, default=60.0, help="HTTP timeout seconds")
    parser.add_argument("--dry-run", action="store_true", help="Resolve sources but do not ingest")
    return parser.parse_args()


def _dataset_summary(
    *,
    dataset: DatasetConfig,
    status: str,
    resolved: ResolvedSource | None,
    records: int,
    schema_validation: dict[str, Any] | None,
    key_field: str | None,
    error: str | None,
    error_code: str | None = None,
    epoch_schedule: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    freshness = (
        summarize_uprn_dataset_freshness(
            dataset_id=dataset.dataset_id,
            resolved_release=resolved.resolved_release if resolved else None,
            resolved_source_url=resolved.resolved_source_url if resolved else None,
            schedule=epoch_schedule,
        )
        if resolved
        else None
    )
    return {
        "id": dataset.dataset_id,
        "kind": dataset.dataset_kind,
        "title": dataset.title,
        "keyType": dataset.key_type,
        "derivationMode": dataset.derivation_mode,
        "priority": dataset.priority,
        "release": dataset.release,
        "resolvedRelease": resolved.resolved_release if resolved else None,
        "status": status,
        "records": records,
        "keyField": key_field,
        "resolverType": dataset.resolver.resolver_type,
        "sourceFormat": resolved.source_format if resolved else None,
        "sourcePath": _display_path(resolved.source_path) if resolved else None,
        "resolvedSourceUrl": resolved.resolved_source_url if resolved else None,
        "sourceSha256": (
            _sha256_file(resolved.source_path)
            if resolved and resolved.source_path.exists()
            else None
        ),
        "retrievedAt": resolved.retrieved_at if resolved else None,
        "schemaValidation": schema_validation,
        "freshness": freshness,
        "errorCode": error_code,
        "error": error,
    }


def _index_health(
    products: list[dict[str, Any]],
    support_products: list[dict[str, Any]],
) -> dict[str, Any]:
    def _mode_ready(mode: str) -> bool:
        mode_products = [
            item
            for item in products
            if item.get("kind") == "product" and item.get("derivationMode") == mode
        ]
        expected_key_types = {
            str(item.get("keyType") or "").strip()
            for item in mode_products
            if str(item.get("keyType") or "").strip()
        }
        ingested_key_types = {
            str(item.get("keyType") or "").strip()
            for item in mode_products
            if item.get("status") == "ingested" and str(item.get("keyType") or "").strip()
        }
        return not expected_key_types or expected_key_types <= ingested_key_types

    exact_ready = _mode_ready("exact")
    best_fit_ready = _mode_ready("best_fit")
    support_ready = (
        all(item.get("status") == "ingested" for item in support_products)
        if support_products
        else True
    )
    freshness_lagging = [
        item.get("id")
        for item in products
        if isinstance(item.get("freshness"), dict)
        and item["freshness"].get("status") == "lagging"
    ]
    degraded_reasons = []
    if not exact_ready:
        degraded_reasons.append("exact_products_unavailable")
    if not best_fit_ready:
        degraded_reasons.append("best_fit_products_unavailable")
    if not support_ready:
        degraded_reasons.append("support_datasets_unavailable")
    if freshness_lagging:
        degraded_reasons.append("outdated_addressbase_epochs")
    return {
        "exactReady": exact_ready,
        "bestFitReady": best_fit_ready,
        "supportReady": support_ready,
        "freshnessReady": not freshness_lagging,
        "laggingProducts": freshness_lagging,
        "status": "ready" if not degraded_reasons else "degraded",
        "degradedReasons": degraded_reasons,
    }


def main() -> int:
    args = parse_args()
    sources_path = Path(args.sources).resolve()
    cache_dir = Path(args.cache_dir).resolve()
    index_path = Path(args.index_path).resolve()
    db_name = str(args.db_name)
    raw_root = cache_dir / "raw"

    try:
        file_overrides = _parse_map_args(args.product_file)
        url_overrides = _parse_map_args(args.product_url)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    version, products, support_products = load_manifest(sources_path)
    if not products:
        raise SystemExit("No valid products found in sources manifest")
    epoch_schedule = load_addressbase_epoch_schedule()

    cache_dir.mkdir(parents=True, exist_ok=True)
    raw_root.mkdir(parents=True, exist_ok=True)

    db_path = cache_dir / db_name
    conn: sqlite3.Connection | None = None
    code_references = CodeReferenceStore()
    summary_products: list[dict[str, Any]] = []
    summary_support: list[dict[str, Any]] = []

    if not args.dry_run:
        conn = sqlite3.connect(str(db_path))
        ensure_schema(conn)

    ordered_support = sorted(support_products, key=lambda item: (item.priority, item.dataset_id))
    ordered_products = sorted(products, key=lambda item: (item.priority, item.dataset_id))

    for dataset in [*ordered_support, *ordered_products]:
        status = "skipped"
        error: str | None = None
        error_code: str | None = None
        resolved: ResolvedSource | None = None
        records = 0
        key_field: str | None = None
        schema_validation: dict[str, Any] | None = None
        try:
            resolved = resolve_dataset_source(
                dataset,
                raw_root=raw_root,
                timeout=float(args.timeout),
                file_overrides=file_overrides,
                url_overrides=url_overrides,
            )
            if args.dry_run:
                status = "resolved"
                schema_validation = {
                    "requiredFound": [],
                    "requiredMissing": dataset.required_fields,
                    "optionalFound": [],
                    "unknownFields": [],
                    "status": "dry_run_unvalidated",
                    "schemaFingerprint": None,
                }
            else:
                assert conn is not None
                if dataset.dataset_kind == "support":
                    records, schema_validation = _ingest_support_dataset(
                        conn=conn,
                        dataset=dataset,
                        resolved=resolved,
                        max_rows=args.max_rows,
                        code_references=code_references,
                    )
                else:
                    records, schema_validation, key_field = _ingest_main_dataset(
                        conn=conn,
                        dataset=dataset,
                        resolved=resolved,
                        max_rows=args.max_rows,
                        code_references=code_references,
                    )
                conn.commit()
                status = "ingested"
        except FileNotFoundError as exc:
            if conn is not None:
                conn.rollback()
            status = "error"
            error = str(exc)
            error_code = "RESOLVE_ERROR"
        except requests.RequestException as exc:
            if conn is not None:
                conn.rollback()
            status = "error"
            error = str(exc)
            error_code = "DOWNLOAD_ERROR"
        except ValueError as exc:
            if conn is not None:
                conn.rollback()
            status = "error"
            error = str(exc)
            if "schema drift" in str(exc).lower():
                error_code = "SCHEMA_DRIFT"
            else:
                error_code = "FORMAT_ERROR"
        except Exception as exc:  # pragma: no cover - defensive
            if conn is not None:
                conn.rollback()
            status = "error"
            error = str(exc)
            error_code = "FORMAT_ERROR"

        summary = _dataset_summary(
            dataset=dataset,
            status=status,
            resolved=resolved,
            records=records,
            schema_validation=schema_validation,
            key_field=key_field,
            error=error,
            error_code=error_code,
            epoch_schedule=epoch_schedule,
        )
        if dataset.dataset_kind == "support":
            summary_support.append(summary)
        else:
            summary_products.append(summary)

    if conn is not None:
        conn.close()

    payload = {
        "version": version,
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "cache": {
            "cacheDir": _display_path(cache_dir),
            "dbPath": _display_path(db_path),
            "rawRoot": _display_path(raw_root),
        },
        "health": _index_health(summary_products, summary_support),
        "supportProducts": summary_support,
        "products": summary_products,
    }
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
